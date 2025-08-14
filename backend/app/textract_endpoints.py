from fastapi import APIRouter, File, UploadFile, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import boto3
import json
import tempfile
import os
import uuid
import pandas as pd
import io
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from datetime import datetime
import csv
import asyncio
import time

router = APIRouter()

# AWS Textract client initialization
def get_textract_client():
    """Initialize AWS Textract client with credentials from environment variables"""
    try:
        client = boto3.client(
            'textract',
            aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
            aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'),
            region_name=os.getenv('AWS_REGION', 'ap-south-1')
        )
        return client
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to initialize AWS Textract client: {str(e)}")

# AWS S3 client for document storage (required for async processing)
def get_s3_client():
    """Initialize AWS S3 client for document storage"""
    try:
        client = boto3.client(
            's3',
            aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
            aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'),
            region_name=os.getenv('AWS_REGION', 'ap-south-1')
        )
        return client
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to initialize AWS S3 client: {str(e)}")

# In-memory job storage (in production, use Redis or database)
textract_jobs = {}

class TextractJobResponse(BaseModel):
    job_id: str
    status: str
    message: str

class TextractJobStatus(BaseModel):
    job_id: str
    status: str
    error: Optional[str] = None
    tables: Optional[List[Dict]] = None

class ExtractedTable(BaseModel):
    table_id: int
    headers: List[str]
    rows: List[List[str]]
    page_number: int

class MergeTablesRequest(BaseModel):
    tables: List[Dict[str, Any]]

class MergedDataResponse(BaseModel):
    headers: List[str]
    rows: List[List[str]]
    total_rows: int
    source_tables: int

@router.post("/start-analysis", response_model=TextractJobResponse)
async def start_textract_analysis(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...)
):
    """Start asynchronous AWS Textract analysis for table extraction"""
    
    if file.content_type != 'application/pdf':
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
    
    # Generate unique job ID
    job_id = str(uuid.uuid4())
    
    try:
        # Read file content
        file_content = await file.read()
        
        # Get S3 bucket name from environment (create bucket if needed)
        s3_bucket = os.getenv('AWS_S3_BUCKET', 'pandiver-textract-documents')
        s3_key = f"documents/{job_id}/{file.filename}"
        
        # Upload to S3 for async processing
        s3_client = get_s3_client()
        
        try:
            s3_client.head_bucket(Bucket=s3_bucket)
        except:
            # Create bucket if it doesn't exist
            s3_client.create_bucket(
                Bucket=s3_bucket,
                CreateBucketConfiguration={'LocationConstraint': os.getenv('AWS_REGION', 'ap-south-1')}
            )
        
        s3_client.put_object(
            Bucket=s3_bucket,
            Key=s3_key,
            Body=file_content,
            ContentType=file.content_type
        )
        
        # Start Textract async analysis
        textract_client = get_textract_client()
        
        response = textract_client.start_document_analysis(
            DocumentLocation={
                'S3Object': {
                    'Bucket': s3_bucket,
                    'Name': s3_key
                }
            },
            FeatureTypes=['TABLES'],
            JobTag=f"pandiver-{job_id}"
        )
        
        # Store job information
        textract_jobs[job_id] = {
            'textract_job_id': response['JobId'],
            'status': 'IN_PROGRESS',
            's3_bucket': s3_bucket,
            's3_key': s3_key,
            'original_filename': file.filename,
            'started_at': datetime.now().isoformat(),
            'tables': None,
            'error': None
        }
        
        return TextractJobResponse(
            job_id=job_id,
            status='IN_PROGRESS',
            message='Textract analysis started successfully'
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to start Textract analysis: {str(e)}")

@router.get("/job-status/{job_id}", response_model=TextractJobStatus)
async def get_job_status(job_id: str):
    """Get the status of a Textract analysis job"""
    
    if job_id not in textract_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job_info = textract_jobs[job_id]
    
    try:
        textract_client = get_textract_client()
        
        # Get job status from Textract
        response = textract_client.get_document_analysis(
            JobId=job_info['textract_job_id']
        )
        
        status = response['JobStatus']
        job_info['status'] = status
        
        if status == 'FAILED':
            job_info['error'] = response.get('StatusMessage', 'Analysis failed')
            
        return TextractJobStatus(
            job_id=job_id,
            status=status,
            error=job_info.get('error'),
            tables=job_info.get('tables')
        )
        
    except Exception as e:
        job_info['status'] = 'FAILED'
        job_info['error'] = str(e)
        
        return TextractJobStatus(
            job_id=job_id,
            status='FAILED',
            error=str(e)
        )

@router.get("/process-results/{job_id}")
async def process_textract_results(job_id: str):
    """Process and extract tables from completed Textract analysis"""
    
    if job_id not in textract_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job_info = textract_jobs[job_id]
    
    if job_info['status'] != 'SUCCEEDED':
        raise HTTPException(status_code=400, detail=f"Job status is {job_info['status']}, cannot process results")
    
    try:
        textract_client = get_textract_client()
        
        # Get all pages of results
        all_blocks = []
        next_token = None
        
        while True:
            if next_token:
                response = textract_client.get_document_analysis(
                    JobId=job_info['textract_job_id'],
                    NextToken=next_token
                )
            else:
                response = textract_client.get_document_analysis(
                    JobId=job_info['textract_job_id']
                )
            
            all_blocks.extend(response['Blocks'])
            
            next_token = response.get('NextToken')
            if not next_token:
                break
        
        # Extract tables from blocks
        tables = extract_tables_from_blocks(all_blocks)
        
        # Store extracted tables
        job_info['tables'] = tables
        
        return {
            'success': True,
            'tables': tables,
            'total_tables': len(tables)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process results: {str(e)}")

def extract_tables_from_blocks(blocks):
    """Extract table data from Textract blocks"""
    
    # Create a mapping of block IDs to blocks
    block_map = {block['Id']: block for block in blocks}
    
    # Find all table blocks
    table_blocks = [block for block in blocks if block['BlockType'] == 'TABLE']
    
    extracted_tables = []
    
    for table_idx, table_block in enumerate(table_blocks):
        # Get page number
        page_number = table_block.get('Page', 1) - 1  # Convert to 0-based
        
        # Extract table structure
        table_data = extract_table_data(table_block, block_map)
        
        if table_data and len(table_data) > 0:
            # First row as headers, rest as data
            headers = table_data[0] if table_data else []
            rows = table_data[1:] if len(table_data) > 1 else []
            
            extracted_tables.append({
                'table_id': table_idx + 1,
                'headers': headers,
                'rows': rows,
                'page_number': page_number
            })
    
    return extracted_tables

def extract_table_data(table_block, block_map):
    """Extract table data from a single table block"""
    
    if 'Relationships' not in table_block:
        return []
    
    # Find CHILD relationships to get cells
    cell_blocks = []
    for relationship in table_block['Relationships']:
        if relationship['Type'] == 'CHILD':
            for cell_id in relationship['Ids']:
                if cell_id in block_map:
                    cell_block = block_map[cell_id]
                    if cell_block['BlockType'] == 'CELL':
                        cell_blocks.append(cell_block)
    
    # Organize cells by row and column
    table_data = {}
    max_row = 0
    max_col = 0
    
    for cell in cell_blocks:
        row_index = cell.get('RowIndex', 1) - 1  # Convert to 0-based
        col_index = cell.get('ColumnIndex', 1) - 1  # Convert to 0-based
        
        max_row = max(max_row, row_index)
        max_col = max(max_col, col_index)
        
        # Extract text from cell
        cell_text = extract_cell_text(cell, block_map)
        
        if row_index not in table_data:
            table_data[row_index] = {}
        table_data[row_index][col_index] = cell_text
    
    # Convert to 2D array
    result = []
    for row_idx in range(max_row + 1):
        row = []
        for col_idx in range(max_col + 1):
            cell_value = table_data.get(row_idx, {}).get(col_idx, '')
            row.append(cell_value)
        result.append(row)
    
    return result

def extract_cell_text(cell_block, block_map):
    """Extract text content from a cell block"""
    
    if 'Relationships' not in cell_block:
        return ''
    
    text_parts = []
    
    for relationship in cell_block['Relationships']:
        if relationship['Type'] == 'CHILD':
            for word_id in relationship['Ids']:
                if word_id in block_map:
                    word_block = block_map[word_id]
                    if word_block['BlockType'] == 'WORD':
                        text_parts.append(word_block.get('Text', ''))
    
    return ' '.join(text_parts).strip()

@router.post("/merge-tables", response_model=MergedDataResponse)
async def merge_tables(request: MergeTablesRequest):
    """Merge multiple extracted tables into a single dataset"""
    
    if not request.tables:
        raise HTTPException(status_code=400, detail="No tables provided for merging")
    
    try:
        # Find the most common header structure
        all_headers = []
        for table in request.tables:
            if table.get('headers'):
                all_headers.append(table['headers'])
        
        if not all_headers:
            raise HTTPException(status_code=400, detail="No headers found in tables")
        
        # Use the longest header set or most common one
        merged_headers = max(all_headers, key=len)
        
        # Merge all rows
        merged_rows = []
        
        for table in request.tables:
            table_headers = table.get('headers', [])
            table_rows = table.get('rows', [])
            
            # Create header mapping
            header_mapping = {}
            for i, header in enumerate(table_headers):
                if header in merged_headers:
                    header_mapping[i] = merged_headers.index(header)
            
            # Add rows with proper column mapping
            for row in table_rows:
                merged_row = [''] * len(merged_headers)
                for old_col_idx, cell_value in enumerate(row):
                    if old_col_idx in header_mapping:
                        new_col_idx = header_mapping[old_col_idx]
                        merged_row[new_col_idx] = cell_value
                merged_rows.append(merged_row)
        
        return MergedDataResponse(
            headers=merged_headers,
            rows=merged_rows,
            total_rows=len(merged_rows),
            source_tables=len(request.tables)
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to merge tables: {str(e)}")

@router.post("/export/{format}")
async def export_data(format: str, data: dict):
    """Export table data in various formats - supports both single tables and multiple tables"""
    
    supported_formats = ['csv', 'xlsx', 'json', 'xml', 'txt']
    if format not in supported_formats:
        raise HTTPException(status_code=400, detail=f"Unsupported format. Supported: {supported_formats}")
    
    # Check if this is multiple tables or merged data
    tables = data.get('tables', [])
    
    if tables:
        # Multiple tables - create separate files
        return export_multiple_tables(format, tables)
    else:
        # Single merged data
        headers = data.get('headers', [])
        rows = data.get('rows', [])
        
        if not headers or not rows:
            raise HTTPException(status_code=400, detail="No data provided for export")
        
        try:
            if format == 'csv':
                return export_csv(headers, rows)
            elif format == 'xlsx':
                return export_excel(headers, rows)
            elif format == 'json':
                return export_json(headers, rows, data)
            elif format == 'xml':
                return export_xml(headers, rows, data)
            elif format == 'txt':
                return export_txt(headers, rows, data)
                
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to export {format}: {str(e)}")

def export_multiple_tables(format: str, tables: List[Dict]):
    """Export multiple tables as a zip file containing separate files for each table"""
    import zipfile
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for table_idx, table in enumerate(tables):
            table_id = table.get('table_id', table_idx + 1)
            page_number = table.get('page_number', 0) + 1  # Convert to 1-based
            headers = table.get('headers', [])
            rows = table.get('rows', [])
            
            if not headers or not rows:
                continue
                
            filename = f"table_{table_id}_page_{page_number}"
            
            if format == 'csv':
                content = generate_csv_content(headers, rows)
                zip_file.writestr(f"{filename}.csv", content)
            elif format == 'xlsx':
                content = generate_excel_content(headers, rows)
                zip_file.writestr(f"{filename}.xlsx", content)
            elif format == 'json':
                content = generate_json_content(headers, rows, table)
                zip_file.writestr(f"{filename}.json", content)
            elif format == 'xml':
                content = generate_xml_content(headers, rows, table)
                zip_file.writestr(f"{filename}.xml", content)
            elif format == 'txt':
                content = generate_txt_content(headers, rows, table)
                zip_file.writestr(f"{filename}.txt", content)
    
    zip_buffer.seek(0)
    
    return StreamingResponse(
        io.BytesIO(zip_buffer.getvalue()),
        media_type='application/zip',
        headers={'Content-Disposition': f'attachment; filename=bank-statement-tables.zip'}
    )

def generate_csv_content(headers, rows):
    """Generate CSV content as string"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()

def generate_excel_content(headers, rows):
    """Generate Excel content as bytes"""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Table Data"
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, cell_value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    wb.save(output)
    return output.getvalue()

def generate_json_content(headers, rows, table_info):
    """Generate JSON content as string"""
    data = {
        'table_id': table_info.get('table_id', 1),
        'page_number': table_info.get('page_number', 0) + 1,
        'headers': headers,
        'data': [
            {headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))}
            for row in rows
        ],
        'metadata': {
            'total_rows': len(rows),
            'export_date': datetime.now().isoformat(),
            'export_type': 'aws_textract_table'
        }
    }
    return json.dumps(data, indent=2, ensure_ascii=False)

def generate_xml_content(headers, rows, table_info):
    """Generate XML content as string"""
    root = ET.Element('table')
    
    # Metadata
    metadata = ET.SubElement(root, 'metadata')
    ET.SubElement(metadata, 'table_id').text = str(table_info.get('table_id', 1))
    ET.SubElement(metadata, 'page_number').text = str(table_info.get('page_number', 0) + 1)
    ET.SubElement(metadata, 'total_rows').text = str(len(rows))
    ET.SubElement(metadata, 'export_date').text = datetime.now().isoformat()
    
    # Data
    data_element = ET.SubElement(root, 'data')
    
    for idx, row in enumerate(rows):
        row_element = ET.SubElement(data_element, 'row', id=str(idx + 1))
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else ''
            clean_header = re.sub(r'[^a-zA-Z0-9]', '_', header.lower())
            ET.SubElement(row_element, clean_header).text = str(value)
    
    return ET.tostring(root, encoding='unicode')

def generate_txt_content(headers, rows, table_info):
    """Generate TXT content as string"""
    output = io.StringIO()
    
    output.write(f'Table {table_info.get("table_id", 1)} - Page {table_info.get("page_number", 0) + 1}\n')
    output.write('=' * 50 + '\n\n')
    
    # Headers
    output.write(' | '.join(headers) + '\n')
    output.write('-' * (len(' | '.join(headers))) + '\n')
    
    # Rows
    for row in rows:
        output.write(' | '.join(str(cell) for cell in row) + '\n')
    
    output.write(f'\nTotal Rows: {len(rows)}\n')
    
    return output.getvalue()

def export_csv(headers, rows):
    """Export data as CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write headers
    writer.writerow(headers)
    
    # Write rows
    for row in rows:
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename=bank-statement-textract.csv'}
    )

def export_excel(headers, rows):
    """Export data as Excel file"""
    output = io.BytesIO()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Transactions"
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, cell_value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=bank-statement-textract.xlsx'}
    )

def export_json(headers, rows, data):
    """Export data as JSON"""
    export_data = {
        'headers': headers,
        'transactions': [
            {headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))}
            for row in rows
        ],
        'metadata': {
            'total_transactions': len(rows),
            'source_tables': data.get('source_tables', 0),
            'export_date': datetime.now().isoformat(),
            'export_type': 'aws_textract'
        }
    }
    
    json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
    
    return StreamingResponse(
        io.BytesIO(json_str.encode('utf-8')),
        media_type='application/json',
        headers={'Content-Disposition': 'attachment; filename=bank-statement-textract.json'}
    )

def export_xml(headers, rows, data):
    """Export data as XML"""
    root = ET.Element('bank_statement')
    
    # Metadata
    metadata = ET.SubElement(root, 'metadata')
    ET.SubElement(metadata, 'total_transactions').text = str(len(rows))
    ET.SubElement(metadata, 'source_tables').text = str(data.get('source_tables', 0))
    ET.SubElement(metadata, 'export_date').text = datetime.now().isoformat()
    ET.SubElement(metadata, 'export_type').text = 'aws_textract'
    
    # Transactions
    transactions = ET.SubElement(root, 'transactions')
    
    for idx, row in enumerate(rows):
        transaction = ET.SubElement(transactions, 'transaction', id=str(idx + 1))
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else ''
            clean_header = re.sub(r'[^a-zA-Z0-9]', '_', header.lower())
            ET.SubElement(transaction, clean_header).text = str(value)
    
    xml_str = ET.tostring(root, encoding='unicode')
    
    return StreamingResponse(
        io.BytesIO(xml_str.encode('utf-8')),
        media_type='application/xml',
        headers={'Content-Disposition': 'attachment; filename=bank-statement-textract.xml'}
    )

def export_txt(headers, rows, data):
    """Export data as text file"""
    output = io.StringIO()
    
    # Header
    output.write('Bank Statement Transactions - AWS Textract\n')
    output.write('=' * 50 + '\n\n')
    
    # Metadata
    output.write(f'Total Transactions: {len(rows)}\n')
    output.write(f'Source Tables: {data.get("source_tables", 0)}\n')
    output.write(f'Export Date: {datetime.now().isoformat()}\n\n')
    
    # Headers
    output.write(' | '.join(headers) + '\n')
    output.write('-' * (len(' | '.join(headers))) + '\n')
    
    # Rows
    for row in rows:
        output.write(' | '.join(str(cell) for cell in row) + '\n')
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type='text/plain',
        headers={'Content-Disposition': 'attachment; filename=bank-statement-textract.txt'}
    )

# Cleanup endpoint to remove old jobs (optional)
@router.delete("/cleanup-jobs")
async def cleanup_old_jobs():
    """Remove jobs older than 24 hours"""
    
    current_time = datetime.now()
    jobs_to_remove = []
    
    for job_id, job_info in textract_jobs.items():
        started_at = datetime.fromisoformat(job_info['started_at'])
        if (current_time - started_at).total_seconds() > 86400:  # 24 hours
            jobs_to_remove.append(job_id)
    
    for job_id in jobs_to_remove:
        del textract_jobs[job_id]
    
    return {
        'message': f'Cleaned up {len(jobs_to_remove)} old jobs',
        'removed_jobs': len(jobs_to_remove)
    }