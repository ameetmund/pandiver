from fastapi import FastAPI, File, UploadFile, HTTPException, Request, Depends, status, Body, Form
import tempfile
import uuid
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import pdfplumber
import io
import pandas as pd
import openpyxl
import json
import re
import collections
# from pdf2image import convert_from_bytes  # Optional dependency for OCR
# from doctr.io import DocumentFile  # Optional dependency for OCR
# from doctr.models import ocr_predictor  # Optional dependency for OCR
from typing import List, Dict, Any, Optional
from pydantic import BaseModel, EmailStr
import jwt
from passlib.context import CryptContext
from datetime import datetime, timedelta
import uuid
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, Session
from .models import Base, User as UserModel, UserTable, ApiKey, ApiUsage, PDFSplitterJob, PDFTranslationJob
# from .tasks import parse_statement  # Optional celery dependency
from .pdf_processor import get_pdf_info, extract_all_words, get_all_page_rows, analyze_row_structure
# from .export_utils import StatementExporter  # Optional dependency
from .bank_parsers import bank_parser_manager
import unicodedata
import sys
import os
# Add the backend directory to the path to import our enhanced parsers
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from pattern_rule_applicator import PatternRuleApplicator, ExtractionResult
from enhanced_bank_parser_v2 import EnhancedTransactionDetector, PatternRule, EnhancedPatternRuleManager

# Import extraction endpoints
from .manual_endpoints import router as manual_router
from .user_controlled_endpoints import router as user_controlled_router
from .textract_endpoints import router as textract_router

# Import Azure Document Intelligence endpoints
from .azure_di_endpoints import router as azure_di_router

# Import new API endpoints
from .api.endpoints import router as api_router
from .api.webhooks import router as webhooks_router
from .api.watch_folder import router as watch_folder_router
from .api_endpoints import router as intelligent_data_router

# Import new PDF features
from .pdf_splitter_endpoints import router as pdf_splitter_router
from .pdf_splitter_api_endpoints import router as pdf_splitter_api_router
from .pdf_translation_endpoints import router as pdf_translation_router
from .pdf_translator_api_endpoints import router as pdf_translator_api_router

app = FastAPI(
    title="Pandiver PDF Processing API",
    description="Advanced PDF processing and transaction extraction API with support for single/bulk processing, webhooks, and watch folders",
    version="2.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# Include existing extraction routes
app.include_router(manual_router, prefix="/manual", tags=["Manual Bank Statement Parser"])
app.include_router(user_controlled_router, prefix="/user-controlled", tags=["User Controlled Column Detection"])
app.include_router(textract_router, prefix="/textract", tags=["AWS Textract Bank Statement Parser"])

# Include Azure Document Intelligence routes
app.include_router(azure_di_router, prefix="/azure-di", tags=["Azure Document Intelligence - Intelligent Tables & Smart Key-Value"])

# Include new API routes
app.include_router(api_router, prefix="/api/v1", tags=["API - File Processing"])
app.include_router(webhooks_router, prefix="/api/v1", tags=["API - Webhooks"])
app.include_router(watch_folder_router, prefix="/api/v1", tags=["API - Watch Folders"])
app.include_router(intelligent_data_router, tags=["Intelligent Data Parser API"])

# Include new PDF feature routes
app.include_router(pdf_splitter_router, tags=["PDF Splitter"])
app.include_router(pdf_splitter_api_router, tags=["PDF Splitter API"])
app.include_router(pdf_translation_router, tags=["PDF Translator"])
app.include_router(pdf_translator_api_router, tags=["PDF Translator API"])

# Database setup
SQLALCHEMY_DATABASE_URL = 'sqlite:///./pandiver.db'
engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base.metadata.create_all(bind=engine)

# Security configuration
SECRET_KEY = "09af8c2e8b3a47f19c6d5e7a8b2c4d6f9e1a3b5c7d9f0e2a4b6c8d0f1e3a5b7c9d1f3e5a7b9c1d3f5e7a9b1d3f5e7a9b"  # In production, use environment variable
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Security scheme
security = HTTPBearer()

# In-memory user storage (replace with database in production)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_db_session():
    """Helper function for background tasks to get DB session"""
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# CORS setup
origins = [
    "http://localhost:3000",  # Next.js default
    "http://127.0.0.1:3000"
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Pydantic models
class UserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: int  # Changed from str to int
    name: str
    email: str
    created_at: datetime

class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

# Authentication helper functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = int(payload.get("sub"))
        if user_id is None:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
        return user_id
    except jwt.PyJWTError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

def get_current_user(user_id: str = Depends(verify_token), db: Session = Depends(get_db)):
    try:
        user_id_int = int(user_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    
    user = db.query(UserModel).filter(UserModel.id == user_id_int).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

# Authentication endpoints

@app.post("/auth/signup", response_model=Token)
async def signup(user_data: UserCreate, db: Session = Depends(get_db)):
    existing_user = db.query(UserModel).filter(UserModel.email == user_data.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    hashed_password = get_password_hash(user_data.password)
    user = UserModel(
        name=user_data.name,
        email=user_data.email,
        hashed_password=hashed_password,
        created_at=datetime.utcnow()
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    access_token = create_access_token({"sub": str(user.id)})
    user_response = UserResponse(
        id=user.id,
        name=user.name,
        email=user.email,
        created_at=user.created_at
    )
    return Token(access_token=access_token, token_type="bearer", user=user_response)

@app.post("/auth/login", response_model=Token)
async def login(user_credentials: UserLogin, db: Session = Depends(get_db)):
    user = db.query(UserModel).filter(UserModel.email == user_credentials.email).first()
    if not user or not verify_password(user_credentials.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    access_token = create_access_token({"sub": str(user.id)})
    user_response = UserResponse(
        id=user.id,
        name=user.name,
        email=user.email,
        created_at=user.created_at
    )
    return Token(access_token=access_token, token_type="bearer", user=user_response)

@app.get("/auth/me", response_model=UserResponse)
async def get_current_user_info(current_user: UserModel = Depends(get_current_user)):
    return UserResponse(
        id=current_user.id,
        name=current_user.name,
        email=current_user.email,
        created_at=current_user.created_at
    )

# API Key management endpoints
class ApiKeyCreate(BaseModel):
    key_name: str

class ApiKeyResponse(BaseModel):
    id: int
    key_name: str
    api_key: str
    is_active: bool
    created_at: datetime
    last_used_at: Optional[datetime]

class ApiUsageResponse(BaseModel):
    id: int
    endpoint: str
    job_id: str
    status: str
    file_count: int
    processing_time: Optional[float]
    created_at: datetime
    completed_at: Optional[datetime]

@app.post("/auth/api-keys", response_model=ApiKeyResponse)
async def create_api_key(
    api_key_data: ApiKeyCreate, 
    current_user: UserModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new API key for the current user"""
    # Generate API key
    import secrets
    api_key = f"pd_{secrets.token_urlsafe(32)}"
    
    # Create new API key record
    new_api_key = ApiKey(
        user_id=current_user.id,
        key_name=api_key_data.key_name,
        api_key=api_key,
        is_active=True,
        created_at=datetime.utcnow()
    )
    
    db.add(new_api_key)
    db.commit()
    db.refresh(new_api_key)
    
    return ApiKeyResponse(
        id=new_api_key.id,
        key_name=new_api_key.key_name,
        api_key=new_api_key.api_key,
        is_active=new_api_key.is_active,
        created_at=new_api_key.created_at,
        last_used_at=new_api_key.last_used_at
    )

@app.get("/auth/api-keys", response_model=List[ApiKeyResponse])
async def get_api_keys(
    current_user: UserModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all API keys for the current user"""
    print(f"DEBUG: API Keys - User ID: {current_user.id}, User email: {current_user.email}")
    api_keys = db.query(ApiKey).filter(ApiKey.user_id == current_user.id).all()
    print(f"DEBUG: API Keys - Found {len(api_keys)} keys for user {current_user.id}")
    
    return [
        ApiKeyResponse(
            id=key.id,
            key_name=key.key_name,
            api_key=key.api_key,
            is_active=key.is_active,
            created_at=key.created_at,
            last_used_at=key.last_used_at
        ) for key in api_keys
    ]

@app.delete("/auth/api-keys/{key_id}")
async def delete_api_key(
    key_id: int,
    current_user: UserModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete an API key"""
    api_key = db.query(ApiKey).filter(
        ApiKey.id == key_id, 
        ApiKey.user_id == current_user.id
    ).first()
    
    if not api_key:
        raise HTTPException(status_code=404, detail="API key not found")
    
    db.delete(api_key)
    db.commit()
    
    return {"message": "API key deleted successfully"}

@app.get("/auth/usage", response_model=List[ApiUsageResponse])
async def get_api_usage(
    current_user: UserModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get API usage history for the current user"""
    usage_records = db.query(ApiUsage).filter(
        ApiUsage.user_id == current_user.id
    ).order_by(ApiUsage.created_at.desc()).limit(100).all()
    
    return [
        ApiUsageResponse(
            id=usage.id,
            endpoint=usage.endpoint,
            job_id=usage.job_id,
            status=usage.status,
            file_count=usage.file_count,
            processing_time=usage.processing_time,
            created_at=usage.created_at,
            completed_at=usage.completed_at
        ) for usage in usage_records
    ]

@app.post("/statement")
async def upload_statement(
    file: UploadFile = File(...),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Upload PDF bank statement for processing.
    
    Saves file to /tmp and enqueues background parsing job.
    Returns job_id for tracking progress.
    """
    # Validate file type
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400, 
            detail="Only PDF files are allowed"
        )
    
    try:
        # Generate unique filename and save to /tmp
        job_id = str(uuid.uuid4())
        file_extension = ".pdf"
        temp_filename = f"statement_{job_id}{file_extension}"
        temp_filepath = os.path.join(tempfile.gettempdir(), temp_filename)
        
        # Save uploaded file
        with open(temp_filepath, "wb") as temp_file:
            content = await file.read()
            temp_file.write(content)
        
        # Enqueue background parsing job
        task = parse_statement.delay(temp_filepath)
        
        return {
            "job_id": task.id,
            "message": "Statement uploaded successfully. Processing started.",
            "status": "queued"
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error processing upload: {str(e)}"
        )

@app.get("/statement/{job_id}/status")
async def get_statement_status(
    job_id: str,
    current_user: UserModel = Depends(get_current_user)
):
    """
    Get status of statement parsing job.
    """
    try:
        from celery.result import AsyncResult
        task_result = AsyncResult(job_id)
        
        if task_result.state == 'PENDING':
            response = {
                'job_id': job_id,
                'state': 'PENDING',
                'message': 'Task is waiting to be processed',
                'progress': 0
            }
        elif task_result.state == 'PROGRESS':
            response = {
                'job_id': job_id,
                'state': 'PROGRESS',
                'message': task_result.info.get('message', ''),
                'progress': task_result.info.get('progress', 0)
            }
        elif task_result.state == 'SUCCESS':
            response = {
                'job_id': job_id,
                'state': 'SUCCESS',
                'result': task_result.result,
                'progress': 100
            }
        else:  # FAILURE
            response = {
                'job_id': job_id,
                'state': 'FAILURE',
                'message': str(task_result.info),
                'progress': 0
            }
        
        return response
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error checking job status: {str(e)}"
        )

@app.post("/statement/analyze")
async def analyze_pdf_structure(
    file: UploadFile = File(...),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Quick analysis of PDF structure (digital vs scanned pages).
    For testing the digital/scanned detector.
    """
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400, 
            detail="Only PDF files are allowed"
        )
    
    try:
        # Read file content
        content = await file.read()
        
        # Save temporarily for analysis
        temp_filename = f"analyze_{uuid.uuid4()}.pdf"
        temp_filepath = os.path.join(tempfile.gettempdir(), temp_filename)
        
        with open(temp_filepath, "wb") as temp_file:
            temp_file.write(content)
        
        # Analyze PDF
        pdf_info = get_pdf_info(temp_filepath)
        
        # Clean up temp file
        os.unlink(temp_filepath)
        
        return {
            "filename": file.filename,
            "analysis": pdf_info,
            "message": f"Analysis complete: {pdf_info['digital_pages']} digital pages, {pdf_info['scanned_pages']} scanned pages"
        }
        
    except Exception as e:
        # Clean up temp file if it exists
        if 'temp_filepath' in locals() and os.path.exists(temp_filepath):
            os.unlink(temp_filepath)
        
        raise HTTPException(
            status_code=500,
            detail=f"Error analyzing PDF: {str(e)}"
        )

@app.post("/statement/extract-words")
async def extract_words_from_pdf(
    file: UploadFile = File(...),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Extract words from PDF for testing word extraction capabilities.
    Shows both digital PDF word extraction and OCR for scanned pages.
    """
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400, 
            detail="Only PDF files are allowed"
        )
    
    try:
        # Read file content
        content = await file.read()
        
        # Save temporarily for analysis
        temp_filename = f"extract_{uuid.uuid4()}.pdf"
        temp_filepath = os.path.join(tempfile.gettempdir(), temp_filename)
        
        with open(temp_filepath, "wb") as temp_file:
            temp_file.write(content)
        
        # Extract all words
        all_words = extract_all_words(temp_filepath)
        
        # Clean up temp file
        os.unlink(temp_filepath)
        
        # Group words by page for easier analysis
        words_by_page = {}
        for word in all_words:
            page_num = word.get('page', 1)
            if page_num not in words_by_page:
                words_by_page[page_num] = []
            words_by_page[page_num].append(word)
        
        return {
            "filename": file.filename,
            "total_words": len(all_words),
            "pages_processed": len(words_by_page),
            "words_by_page": {
                str(page): {
                    "word_count": len(words),
                    "sample_words": words[:10]  # First 10 words as sample
                }
                for page, words in words_by_page.items()
            },
            "message": f"Word extraction complete: {len(all_words)} words from {len(words_by_page)} pages"
        }
        
    except Exception as e:
        # Clean up temp file if it exists
        if 'temp_filepath' in locals() and os.path.exists(temp_filepath):
            os.unlink(temp_filepath)
        
        raise HTTPException(
            status_code=500,
            detail=f"Error extracting words: {str(e)}"
        )

@app.post("/statement/cluster-rows")
async def cluster_pdf_rows(
    file: UploadFile = File(...),
    tolerance: float = 2.0,
    current_user: UserModel = Depends(get_current_user)
):
    """
    Extract words and cluster them into rows for testing row clustering.
    Shows how words are grouped into horizontal rows based on Y coordinates.
    """
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400, 
            detail="Only PDF files are allowed"
        )
    
    try:
        # Read file content
        content = await file.read()
        
        # Save temporarily for analysis
        temp_filename = f"cluster_{uuid.uuid4()}.pdf"
        temp_filepath = os.path.join(tempfile.gettempdir(), temp_filename)
        
        with open(temp_filepath, "wb") as temp_file:
            temp_file.write(content)
        
        # Extract all words
        all_words = extract_all_words(temp_filepath)
        
        # Cluster into rows
        all_page_rows = get_all_page_rows(all_words, tolerance=tolerance)
        
        # Clean up temp file
        os.unlink(temp_filepath)
        
        # Analyze structure for each page
        page_analyses = {}
        total_rows = 0
        
        for page_num, rows in all_page_rows.items():
            analysis = analyze_row_structure(rows)
            page_analyses[str(page_num)] = analysis
            total_rows += len(rows)
        
        # Sample rows for display (first 3 rows from first page)
        sample_rows = []
        first_page_rows = all_page_rows.get(1, [])
        for i, row in enumerate(first_page_rows[:3]):
            row_text = " ".join(word.get('text', '') for word in row)
            sample_rows.append({
                "row_index": i,
                "word_count": len(row),
                "text": row_text,
                "y_range": {
                    "min_y": min(w.get('y0', 0) for w in row) if row else 0,
                    "max_y": max(w.get('y1', 0) for w in row) if row else 0,
                    "mid_y": sum((w.get('y0', 0) + w.get('y1', 0)) / 2 for w in row) / len(row) if row else 0
                }
            })
        
        return {
            "filename": file.filename,
            "tolerance_used": tolerance,
            "total_words": len(all_words),
            "total_rows": total_rows,
            "pages_processed": len(all_page_rows),
            "page_analyses": page_analyses,
            "sample_rows": sample_rows,
            "message": f"Row clustering complete: {len(all_words)} words clustered into {total_rows} rows across {len(all_page_rows)} pages"
        }
        
    except Exception as e:
        # Clean up temp file if it exists
        if 'temp_filepath' in locals() and os.path.exists(temp_filepath):
            os.unlink(temp_filepath)
        
        raise HTTPException(
            status_code=500,
            detail=f"Error clustering rows: {str(e)}"
        )

@app.get("/")
async def root():
    return {"message": "PDF Text Extraction API"}

# PDF processing functions
def extract_text_blocks(pdf_bytes):
    results = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages):
            blocks = []
            for char in page.extract_words(x_tolerance=1, y_tolerance=1):
                block = {
                    "text": char["text"],
                    "x0": char["x0"],
                    "y0": char["top"],
                    "x1": char["x1"],
                    "y1": char["bottom"]
                }
                blocks.append(block)
            results.append({
                "page": page_num,
                "blocks": blocks
            })
    return results

def detect_table_structure(pdf_bytes):
    """
    Detect table structure and extract columns with smart consolidation across pages
    """
    results = []
    all_tables = []
    
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages):
            # Extract tables from the page
            tables = page.extract_tables()
            
            page_info = {
                "page": page_num,
                "tables": [],
                "detected_columns": []
            }
            
            for table_idx, table in enumerate(tables):
                if table and len(table) > 0:
                    # Process table to identify columns
                    headers = table[0] if table else []
                    data_rows = table[1:] if len(table) > 1 else []
                    
                    # Clean headers (remove None/empty values)
                    clean_headers = []
                    for i, header in enumerate(headers):
                        if header and header.strip():
                            clean_headers.append({
                                "text": header.strip(),
                                "column_index": i,
                                "x_position": i * 100,  # Approximate position
                                "page": page_num,
                                "table_index": table_idx
                            })
                    
                    # Extract column data
                    columns = []
                    for col_idx, header in enumerate(clean_headers):
                        column_data = []
                        for row in data_rows:
                            if col_idx < len(row) and row[col_idx]:
                                column_data.append(row[col_idx].strip())
                        
                        columns.append({
                            "header": header,
                            "data": column_data,
                            "column_index": col_idx,
                            "page": page_num
                        })
                    
                    page_info["tables"].append({
                        "table_index": table_idx,
                        "headers": clean_headers,
                        "columns": columns,
                        "raw_data": table
                    })
                    
                    # Add to all tables for cross-page consolidation
                    all_tables.extend(columns)
            
            results.append(page_info)
    
    # Consolidate columns across pages
    consolidated_columns = consolidate_columns_across_pages(all_tables)
    
    return {
        "pages": results,
        "consolidated_columns": consolidated_columns
    }

def consolidate_columns_across_pages(all_columns: List[Dict]) -> List[Dict]:
    """
    Consolidate columns with similar headers across multiple pages
    """
    if not all_columns:
        return []
    
    # Group columns by similar header names
    column_groups = {}
    
    for column in all_columns:
        header_text = column["header"]["text"].lower().strip()
        
        # Find matching group or create new one
        matched_group = None
        for existing_header in column_groups.keys():
            # Simple similarity check - can be improved with fuzzy matching
            if header_text == existing_header or are_headers_similar(header_text, existing_header):
                matched_group = existing_header
                break
        
        if matched_group:
            column_groups[matched_group].append(column)
        else:
            column_groups[header_text] = [column]
    
    # Consolidate data for each group
    consolidated = []
    for header_name, columns in column_groups.items():
        # Use the first column's header as the primary header
        primary_header = columns[0]["header"]
        
        # Merge all data from all pages
        all_data = []
        pages_found = []
        
        for col in columns:
            all_data.extend(col["data"])
            if col["page"] not in pages_found:
                pages_found.append(col["page"])
        
        # Remove duplicates while preserving order
        unique_data = []
        seen = set()
        for item in all_data:
            if item not in seen:
                unique_data.append(item)
                seen.add(item)
        
        consolidated.append({
            "header": primary_header,
            "data": unique_data,
            "pages": sorted(pages_found),
            "total_rows": len(unique_data),
            "is_consolidated": len(pages_found) > 1
        })
    
    return consolidated

def are_headers_similar(header1: str, header2: str) -> bool:
    """
    Check if two headers are similar enough to be considered the same column
    """
    # Simple similarity check - can be enhanced with more sophisticated algorithms
    header1_words = set(header1.split())
    header2_words = set(header2.split())
    
    if not header1_words or not header2_words:
        return False
    
    # Check if they have significant word overlap
    intersection = header1_words.intersection(header2_words)
    union = header1_words.union(header2_words)
    
    similarity = len(intersection) / len(union) if union else 0
    return similarity > 0.6  # 60% similarity threshold

# PDF processing endpoints
@app.post("/extract")
async def extract_pdf(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF.")
    pdf_bytes = await file.read()
    try:
        blocks = extract_text_blocks(pdf_bytes)
        return JSONResponse(content={"pages": blocks})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/detect-tables")
async def detect_tables(file: UploadFile = File(...)):
    """
    Detect table structure and return consolidated columns for Bank Statement Parser
    """
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF.")
    
    pdf_bytes = await file.read()
    try:
        table_structure = detect_table_structure(pdf_bytes)
        return JSONResponse(content=table_structure)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# @app.post("/export")
async def export_table(request: Request):
    data = await request.json()
    df = pd.DataFrame(data.get("table", []))
    fmt = data.get("format", "xlsx")
    output = io.BytesIO()
    if fmt == "csv":
        df.to_csv(output, index=False)
        output.seek(0)
        return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=table.csv"})
    else:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=table.xlsx"})

@app.post("/extract-text-region")
async def extract_text_region(
    file: UploadFile = File(...),
    page: str = Form(...),
    x: str = Form(...),
    y: str = Form(...),
    width: str = Form(...),
    height: str = Form(...)
):
    try:
        # Parse form data to proper types
        page_num = int(page)
        x_coord = float(x)
        y_coord = float(y)
        width_val = float(width)
        height_val = float(height)
        
        # Log the received coordinates for debugging
        print(f"[Backend] Received coordinates: page={page_num}, x={x_coord}, y={y_coord}, width={width_val}, height={height_val}")
        
        # Read the PDF file
        pdf_bytes = await file.read()
        
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            if page_num >= len(pdf.pages):
                raise HTTPException(status_code=400, detail="Page number out of range")
            
            pdf_page = pdf.pages[page_num]
            
            # Extract text from the specified region
            region = pdf_page.crop((x_coord, y_coord, x_coord + width_val, y_coord + height_val))
            
            # Try to extract as table first for better structure preservation
            tables = region.extract_tables()
            print(f"[Backend] Found {len(tables) if tables else 0} tables in region")
            table_found = False
            if tables and len(tables) > 0:
                # If we found a table, convert it to structured text
                table = tables[0]
                structured_text = []
                for row in table:
                    if row:  # Skip empty rows
                        cleaned_row = [cell.strip() if cell else "" for cell in row if cell is not None]
                        if any(cleaned_row):  # Only add non-empty rows
                            structured_text.append(cleaned_row)
                text = structured_text
                table_found = True
                print(f"[Backend] Extracted table with {len(structured_text)} rows")
            else:
                # Fallback to regular text extraction with better formatting
                text = region.extract_text(x_tolerance=3, y_tolerance=3)
                print(f"[Backend] Extracted text: '{text[:100] if text else 'None'}...'")
            
            print(f"[Backend] Final text type: {type(text)}, length: {len(text) if text else 0}")
            
            return {
                "text": text if text else "",
                "page": page_num,
                "region": {
                    "x": x_coord,
                    "y": y_coord,
                    "width": width_val,
                    "height": height_val
                },
                "table_found": table_found
            }
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error extracting text: {str(e)}")

@app.post("/extract-multi-page-table")
async def extract_multi_page_table(
    file: UploadFile = File(...),
    start_page: str = Form(...),
    end_page: str = Form(...),
    x: str = Form(...),
    y: str = Form(...),
    width: str = Form(...),
    height: str = Form(...)
):
    try:
        # Parse form data
        start_page_num = int(start_page)
        end_page_num = int(end_page)
        x_coord = float(x)
        y_coord = float(y)
        width_val = float(width)
        height_val = float(height)
        
        print(f"[Backend] Multi-page extraction: pages {start_page_num}-{end_page_num}, coords=({x_coord}, {y_coord}, {width_val}, {height_val})")
        
        # Read the PDF file
        pdf_bytes = await file.read()
        
        all_table_data = []
        all_text_data = []
        page_summaries = []
        headers = None
        
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_num in range(start_page_num, min(end_page_num + 1, len(pdf.pages))):
                pdf_page = pdf.pages[page_num]
                region = pdf_page.crop((x_coord, y_coord, x_coord + width_val, y_coord + height_val))
                tables = region.extract_tables()
                if tables and len(tables) > 0:
                    table = tables[0]
                    page_table_data = []
                    for i, row in enumerate(table):
                        if row and any(cell and cell.strip() for cell in row):
                            cleaned_row = [cell.strip() if cell else "" for cell in row]
                            if headers is None and i == 0:
                                headers = cleaned_row
                            elif headers and cleaned_row != headers:
                                page_table_data.append(cleaned_row)
                            elif headers is None:
                                page_table_data.append(cleaned_row)
                    if page_table_data:
                        all_table_data.extend(page_table_data)
                        page_summaries.append({"page": page_num, "type": "table", "rows": len(page_table_data)})
                    else:
                        page_summaries.append({"page": page_num, "type": "empty-table"})
                else:
                    # Fallback to text extraction if no table found
                    text = region.extract_text(x_tolerance=3, y_tolerance=3)
                    if text and text.strip():
                        all_text_data.append({"page": page_num, "text": text})
                        page_summaries.append({"page": page_num, "type": "text", "length": len(text)})
                    else:
                        page_summaries.append({"page": page_num, "type": "empty"})
        return {
            "headers": headers,
            "data": all_table_data,
            "text_fallback": all_text_data,
            "pages_processed": list(range(start_page_num, min(end_page_num + 1, len(pdf.pages)))),
            "total_rows": len(all_table_data),
            "page_summaries": page_summaries
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error extracting multi-page table: {str(e)}")

@app.post("/get-text-spans")
async def get_text_spans(
    file: UploadFile = File(...),
    page: str = Form(...)
):
    """Get all text spans with their positions for column-based selection"""
    try:
        page_num = int(page)
        pdf_bytes = await file.read()
        
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            if page_num >= len(pdf.pages):
                raise HTTPException(status_code=400, detail="Page number out of range")
            
            pdf_page = pdf.pages[page_num]
            
            # Extract individual characters with positions
            chars = pdf_page.chars
            
            # Group characters into words
            words = []
            current_word = {"text": "", "chars": []}
            
            for char in chars:
                # If character is space or we're starting, finish current word
                if char.get('text', '').strip() == '' or len(current_word["chars"]) == 0:
                    if current_word["text"].strip():
                        # Calculate word boundaries
                        x0 = min(c['x0'] for c in current_word["chars"])
                        x1 = max(c['x1'] for c in current_word["chars"])
                        y0 = min(c['y0'] for c in current_word["chars"])
                        y1 = max(c['y1'] for c in current_word["chars"])
                        
                        words.append({
                            "text": current_word["text"],
                            "x0": x0,
                            "y0": y0,
                            "x1": x1,
                            "y1": y1,
                            "width": x1 - x0,
                            "height": y1 - y0
                        })
                    
                    # Start new word
                    current_word = {"text": char.get('text', ''), "chars": [char] if char.get('text', '').strip() else []}
                else:
                    # Add to current word
                    current_word["text"] += char.get('text', '')
                    current_word["chars"].append(char)
            
            # Don't forget the last word
            if current_word["text"].strip():
                x0 = min(c['x0'] for c in current_word["chars"])
                x1 = max(c['x1'] for c in current_word["chars"])
                y0 = min(c['y0'] for c in current_word["chars"])
                y1 = max(c['y1'] for c in current_word["chars"])
                
                words.append({
                    "text": current_word["text"],
                    "x0": x0,
                    "y0": y0,
                    "x1": x1,
                    "y1": y1,
                    "width": x1 - x0,
                    "height": y1 - y0
                })
            
            print(f"[Backend] Found {len(words)} words on page {page_num}")
            
            return {
                "words": words,
                "page": page_num,
                "page_width": pdf_page.width,
                "page_height": pdf_page.height
            }
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting text spans: {str(e)}")

# Smart Bank Statement parsing utilities
def is_digital_pdf(pdf_page):
    """Check if PDF page has extractable text (digital) or needs OCR (scanned)"""
    try:
        text = pdf_page.extract_text()
        return bool(text and text.strip())
    except:
        return False

def get_words_from_digital(pdf_page):
    """Extract words from digital PDF using pdfplumber, robust to key differences."""
    try:
        words = pdf_page.extract_words()
        result = []
        for word in words:
            try:
                y0 = word.get("y0", word.get("top"))
                y1 = word.get("y1", word.get("bottom"))
                if y0 is None or y1 is None:
                    print(f"[Smart Extract] Warning: Word missing y0/y1/top/bottom: {word}")
                    continue
                result.append({
                    "text": word["text"],
                    "x0": word["x0"],
                    "y0": y0,
                    "x1": word["x1"],
                    "y1": y1
                })
            except Exception as e:
                print(f"[Smart Extract] Error processing word: {word}, error: {e}")
        print(f"[Smart Extract] Extracted {len(result)} words from digital PDF page.")
        return result
    except Exception as e:
        print(f"Error extracting words from digital PDF: {e}")
        return []

def get_words_from_scanned(page_image):
    """Extract words from scanned page using DocTR OCR"""
    try:
        # Initialize OCR model (you might want to cache this)
        model = ocr_predictor(pretrained=True)
        
        # Convert PIL image to format expected by DocTR
        doc = DocumentFile.from_images([page_image])
        
        # Perform OCR
        result = model(doc)
        
        words = []
        for page in result.pages:
            for block in page.blocks:
                for line in block.lines:
                    for word in line.words:
                        # Convert relative coordinates to absolute
                        x0 = word.geometry[0][0] * page_image.width
                        y0 = word.geometry[0][1] * page_image.height
                        x1 = word.geometry[1][0] * page_image.width
                        y1 = word.geometry[1][1] * page_image.height
                        
                        words.append({
                            "text": word.value,
                            "x0": x0,
                            "y0": y0,
                            "x1": x1,
                            "y1": y1
                        })
        return words
    except Exception as e:
        print(f"Error extracting words from scanned PDF: {e}")
        return []

def cluster_rows(words, y_tolerance=3):
    """Group words into rows based on Y-coordinate proximity"""
    if not words:
        return []
    
    rows = collections.defaultdict(list)
    for word in sorted(words, key=lambda w: w["y0"]):
        # Group words with similar Y coordinates
        row_key = round(word["y0"] / y_tolerance) * y_tolerance
        rows[row_key].append(word)
    
    # Sort each row by X coordinate and return as list
    return [sorted(row, key=lambda w: w["x0"]) for row in rows.values()]

def detect_headers_and_column_map(words, header_keywords=None):
    """Detect headers and map columns to standard fields using keywords."""
    if header_keywords is None:
        header_keywords = {
            'date': ['date', 'transaction date', 'value date'],
            'description': ['description', 'particulars', 'details', 'narration', 'transaction details'],
            'debit': ['debit', 'withdrawal', 'withdrawals', 'dr', 'debits'],
            'credit': ['credit', 'deposit', 'deposits', 'cr', 'credits'],
            'amount': ['amount', 'amt'],
            'balance': ['balance', 'bal', 'closing balance', 'available balance']
        }
    # Find the row with the most header-like words
    rows = cluster_rows(words, y_tolerance=3)
    for row in rows[:5]:  # Only check first 5 rows
        header_map = {}
        for i, word in enumerate(row):
            text = word['text'].strip().lower()
            for key, variants in header_keywords.items():
                if any(variant in text for variant in variants):
                    header_map[key] = i
        if len(header_map) >= 3:  # At least 3 columns detected
            return header_map, row
    return None, None

def parse_bank_transactions(words):
    """Parse bank statement transactions using header detection and regex patterns."""
    # Enhanced regex for amounts (supports commas, periods, negatives, currency)
    amount_pattern = r'[-+]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?'
    currency_pattern = r'[₹$€£¥]'
    date_pattern = r'\b(\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|\d{2,4}[-/]\d{1,2}[-/]\d{1,2})\b'
    rows = cluster_rows(words, y_tolerance=3)
    header_map, header_row = detect_headers_and_column_map(words)
    transactions = []
    prev_balance = None
    current_transaction = None
    for row in rows:
        row_text = " ".join(word["text"] for word in row)
        # If header detected, skip header row
        if header_row and row == header_row:
            continue
        # If header_map is available, use column positions
        if header_map:
            fields = {}
            for key, idx in header_map.items():
                if idx < len(row):
                    fields[key] = row[idx]["text"].strip()
            # Try to parse date
            date_val = fields.get('date')
            if date_val and re.match(date_pattern, date_val):
                # Parse amounts
                debit = fields.get('debit')
                credit = fields.get('credit')
                amount = fields.get('amount')
                balance = fields.get('balance')
                # Clean and parse amounts
                def parse_amt(val):
                    if not val: return None
                    val = re.sub(currency_pattern, '', val)
                    val = val.replace(',', '').replace(' ', '')
                    try:
                        return float(val)
                    except:
                        return None
                debit_val = parse_amt(debit)
                credit_val = parse_amt(credit)
                amount_val = parse_amt(amount)
                balance_val = parse_amt(balance)
                particulars = fields.get('description', '')
                transactions.append({
                    "Date": date_val,
                    "Particulars": particulars,
                    "Deposits": credit_val or (amount_val if (amount_val and amount_val > 0) else None),
                    "Withdrawals": debit_val or (amount_val if (amount_val and amount_val < 0) else None),
                    "Balance": balance_val
                })
                prev_balance = balance_val
            else:
                # Multi-line description continuation
                if transactions and fields.get('description'):
                    transactions[-1]['Particulars'] += ' ' + fields['description']
            continue
        # Fallback: regex-based parsing
        date_match = re.search(date_pattern, row_text)
        amounts = re.findall(amount_pattern, row_text)
        # Remove currency symbols
        amounts = [re.sub(currency_pattern, '', amt) for amt in amounts]
        # Try to parse floats
        amounts_float = []
        for amt in amounts:
            try:
                amt_clean = amt.replace(',', '').replace(' ', '')
                amounts_float.append(float(amt_clean))
            except:
                continue
        if date_match and amounts_float:
            if current_transaction:
                transactions.append(current_transaction)
            balance = amounts_float[-1]
            transaction_amount = amounts_float[-2] if len(amounts_float) >= 2 else None
            deposits = None
            withdrawals = None
            if transaction_amount and prev_balance is not None:
                if balance > prev_balance:
                    deposits = transaction_amount
                else:
                    withdrawals = transaction_amount
            date_text = date_match.group()
            # Description: all text after date, before first amount
            desc_start = row_text.find(date_text) + len(date_text)
            desc_end = row_text.find(amounts[0]) if amounts else len(row_text)
            particulars = row_text[desc_start:desc_end].strip()
            # Clean up particulars (remove only amounts that are not part of description)
            for amt in amounts:
                if amt in particulars:
                    particulars = particulars.replace(amt, '').strip()
            current_transaction = {
                "Date": date_text,
                "Particulars": particulars,
                "Deposits": deposits,
                "Withdrawals": withdrawals,
                "Balance": balance
            }
            prev_balance = balance
        else:
            # Multi-line description continuation
            if current_transaction and row_text.strip():
                current_transaction["Particulars"] += " " + row_text.strip()
    if current_transaction:
        transactions.append(current_transaction)
    return transactions

@app.post("/smart-extract-bank-statement")
async def smart_extract_bank_statement(file: UploadFile = File(...)):
    """Universal smart extraction for bank statements using OCR and advanced parsing"""
    try:
        print(f"[Smart Extract] Starting universal extraction for file: {file.filename}")
        
        # Use the existing bank_parser_manager (imported at top)
        
        pdf_bytes = await file.read()
        all_words = []
        page_width = 0
        
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            total_pages = len(pdf.pages)
            print(f"[Smart Extract] Processing {total_pages} pages")
            
            for page_num, pdf_page in enumerate(pdf.pages):
                print(f"[Smart Extract] Processing page {page_num + 1}/{total_pages}")
                
                # Get page width for RTL detection
                if page_num == 0:
                    page_width = pdf_page.width
                
                # Check if page is digital or scanned
                if is_digital_pdf(pdf_page):
                    print(f"[Smart Extract] Page {page_num + 1} is digital, using pdfplumber")
                    words = get_words_from_digital(pdf_page)
                else:
                    print(f"[Smart Extract] Page {page_num + 1} is scanned, using OCR")
                    # Convert page to image for OCR
                    page_images = convert_from_bytes(
                        pdf_bytes, 
                        dpi=200, 
                        first_page=page_num + 1, 
                        last_page=page_num + 1
                    )
                    if page_images:
                        words = get_words_from_scanned(page_images[0])
                    else:
                        words = []
                
                print(f"[Smart Extract] Found {len(words)} words on page {page_num + 1}")
                
                # Add page number to each word
                for word in words:
                    word['page'] = page_num + 1
                
                all_words.extend(words)
        
        print(f"[Smart Extract] Total words extracted: {len(all_words)}")
        
        # Use the Bank-Specific Parser Manager for intelligent parsing
        print(f"[Smart Extract] Using Bank-Specific Parser Manager...")
        
        # Convert words to rows format expected by bank parser
        from .universal_parser import words_to_rows
        all_page_rows = words_to_rows(all_words)
        print(f"[Smart Extract] Created {len(all_page_rows)} rows for bank detection")
        
        formatted_transactions = bank_parser_manager.detect_bank_and_parse(all_page_rows)
        
        # If no transactions found, provide helpful error message
        if not formatted_transactions:
            print("[Smart Extract] No transactions found by bank-specific parser")
            return {
                "transactions": [],
                "summary": {
                    "total_transactions": 0,
                    "total_pages": total_pages,
                    "error": "No transactions found in the bank statement. Please try manual mode.",
                    "suggestions": [
                        "The document may not be a standard bank statement format",
                        "Try using the manual extraction mode instead",
                        "Ensure the PDF contains transaction data with dates and amounts"
                    ]
                },
                "status": "no_transactions_found"
            }
        
        print(f"[Smart Extract] Total transactions found: {len(formatted_transactions)}")
        
        # Generate summary from parsed transactions
        summary = {
            "total_transactions": len(formatted_transactions),
            "total_pages": total_pages,
            "date_range": {
                "start": None,
                "end": None
            },
            "total_deposits": 0,
            "total_withdrawals": 0,
            "final_balance": None
        }
        
        if formatted_transactions:
            # Extract summary information from parsed transactions
            dates = []
            total_credits = 0
            total_debits = 0
            final_balance = None
            
            for txn in formatted_transactions:
                # Collect dates
                if txn.get('Date'):
                    dates.append(txn['Date'])
                
                # Sum credits/debits
                if txn.get('Credit') and isinstance(txn['Credit'], (int, float)):
                    total_credits += txn['Credit']
                if txn.get('Debit') and isinstance(txn['Debit'], (int, float)):
                    total_debits += txn['Debit']
                
                # Get final balance (from last transaction)
                if txn.get('Balance') and isinstance(txn['Balance'], (int, float)):
                    final_balance = txn['Balance']
            
            # Set date range
            if dates:
                summary["date_range"]["start"] = dates[0]
                summary["date_range"]["end"] = dates[-1]
            
            summary["total_deposits"] = total_credits
            summary["total_withdrawals"] = total_debits
            summary["final_balance"] = final_balance
        
        return {
            "transactions": formatted_transactions,
            "summary": summary,
            "status": "success"
        }
        
    except Exception as e:
        print(f"[Smart Extract] Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error in smart extraction: {str(e)}")

# @app.post("/export-excel")
async def export_excel(data: dict):
    try:
        headers = data.get("headers", [])
        rows = data.get("rows", [])
        
        if not headers and not rows:
            raise HTTPException(status_code=400, detail="No data to export")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bank Statement Data"
        
        # Add headers if provided
        if headers:
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
                # Make header bold
                ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)
        
        # Add data rows
        start_row = 2 if headers else 1
        for row_idx, row_data in enumerate(rows, start_row):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=bank-statement-data.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting to Excel: {str(e)}")

# Save a table for the current user
@app.post("/save-table")
async def save_table(
    name: str = Body(...),
    data: dict = Body(...),
    current_user: UserModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Save table as JSON string
    table_json = json.dumps(data)
    user_table = UserTable(
        user_id=current_user.id,
        name=name,
        data=table_json,
        created_at=datetime.utcnow()
    )
    db.add(user_table)
    db.commit()
    db.refresh(user_table)
    return {"id": user_table.id, "name": user_table.name, "created_at": user_table.created_at}

# Get all tables for the current user
@app.get("/get-tables")
async def get_tables(
    current_user: UserModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    tables = db.query(UserTable).filter(UserTable.user_id == current_user.id).all()
    result = []
    for t in tables:
        result.append({
            "id": t.id,
            "name": t.name,
            "data": json.loads(t.data),
            "created_at": t.created_at
        })
    return result

class ColumnBoundary(BaseModel):
    name: str
    x0: float
    x1: float

class MultiPageColumnsRequest(BaseModel):
    start_page: int
    end_page: int
    columns: List[ColumnBoundary]

@app.post("/extract-multi-page-columns")
async def extract_multi_page_columns(
    file: UploadFile = File(...),
    start_page: int = Form(...),
    end_page: int = Form(...),
    columns: str = Form(...),  # JSON stringified list of {name, x0, x1}
):
    """
    Extracts text for each user-defined column boundary across a page range.
    columns: JSON stringified list of {name, x0, x1}
    """
    try:
        pdf_bytes = await file.read()
        columns_list = json.loads(columns)
        # Validate columns
        if not isinstance(columns_list, list) or not all('x0' in c and 'x1' in c and 'name' in c for c in columns_list):
            raise HTTPException(status_code=400, detail="Invalid columns format")
        
        results = {col['name']: [] for col in columns_list}
        page_summaries = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_num in range(start_page, min(end_page + 1, len(pdf.pages))):
                pdf_page = pdf.pages[page_num]
                page_summary = {"page": page_num, "columns": []}
                for col in columns_list:
                    region = pdf_page.crop((col['x0'], 0, col['x1'], pdf_page.height))
                    text = region.extract_text(x_tolerance=3, y_tolerance=3)
                    if text and text.strip():
                        lines = [line.strip() for line in text.split('\n') if line.strip()]
                        results[col['name']].extend(lines)
                        page_summary["columns"].append({"name": col['name'], "lines": len(lines)})
                    else:
                        page_summary["columns"].append({"name": col['name'], "lines": 0})
                page_summaries.append(page_summary)
        return {
            "columns": [{"name": name, "data": data} for name, data in results.items()],
            "page_summaries": page_summaries
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error extracting multi-page columns: {str(e)}")

@app.post("/extract-pdf-tables-faithful")
async def extract_pdf_tables_faithful(file: UploadFile = File(...)):
    """
    Extract all tables from a PDF, preserving original headers and column order as in the PDF.
    If headers change across pages, output separate tables for each header set.
    """
    try:
        pdf_bytes = await file.read()
        tables_by_header = {}
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue  # Skip empty or header-only tables
                    header = tuple(cell.strip() if cell else "" for cell in table[0])
                    data_rows = [
                        [cell.strip() if cell else "" for cell in row]
                        for row in table[1:]
                    ]
                    if header in tables_by_header:
                        tables_by_header[header].extend(data_rows)
                    else:
                        tables_by_header[header] = data_rows
        # Prepare output: list of tables, each with headers and rows
        output_tables = []
        for header, rows in tables_by_header.items():
            output_tables.append({
                "headers": list(header),
                "rows": rows
            })
        return {"tables": output_tables}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error extracting tables: {str(e)}")


# @app.post("/export-transactions-excel")
async def export_transactions_excel(
    transactions: List[Dict[str, Any]] = Body(...),
    bank_name: str = Body("Bank"),
    statement_period: str = Body(""),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Export transactions to Excel format.
    
    Args:
        transactions: List of transaction dictionaries
        bank_name: Name of the bank for the header
        statement_period: Statement period string
        
    Returns:
        Excel file as StreamingResponse
    """
    try:
        exporter = StatementExporter()
        
        # Generate Excel data
        excel_data = exporter.to_excel(
            transactions=transactions,
            bank_name=bank_name,
            statement_period=statement_period
        )
        
        # Generate filename
        filename = exporter.get_export_filename(bank_name, 'xlsx', statement_period)
        
        # Create streaming response
        def generate():
            yield excel_data
        
        return StreamingResponse(
            generate(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting to Excel: {str(e)}")


# @app.post("/export-transactions-csv")
async def export_transactions_csv(
    transactions: List[Dict[str, Any]] = Body(...),
    bank_name: str = Body("Bank"), 
    statement_period: str = Body(""),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Export transactions to CSV format.
    
    Args:
        transactions: List of transaction dictionaries
        bank_name: Name of the bank for the header
        statement_period: Statement period string
        
    Returns:
        CSV file as StreamingResponse
    """
    try:
        exporter = StatementExporter()
        
        # Generate CSV data
        csv_data = exporter.to_csv(transactions=transactions)
        
        # Generate filename
        filename = exporter.get_export_filename(bank_name, 'csv', statement_period)
        
        # Create streaming response
        def generate():
            yield csv_data.encode('utf-8')
        
        return StreamingResponse(
            generate(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting to CSV: {str(e)}")


@app.get("/export-sample")
async def export_sample(format_type: str = "excel"):
    """
    Generate sample export for testing purposes.
    
    Args:
        format_type: 'excel' or 'csv'
        
    Returns:
        Sample export file
    """
    try:
        # Sample transaction data
        sample_transactions = [
            {'Date': '01/06/2025', 'Description': 'Opening Balance', 'Debit': '0.00', 'Credit': '0.00', 'Balance': '5000.00'},
            {'Date': '02/06/2025', 'Description': 'ATM Withdrawal - HDFC ATM', 'Debit': '500.00', 'Credit': '0.00', 'Balance': '4500.00'},
            {'Date': '03/06/2025', 'Description': 'Salary Credit - Company XYZ', 'Debit': '0.00', 'Credit': '5000.00', 'Balance': '9500.00'},
            {'Date': '04/06/2025', 'Description': 'Online Purchase - Amazon', 'Debit': '250.75', 'Credit': '0.00', 'Balance': '9249.25'},
            {'Date': '05/06/2025', 'Description': 'Interest Credit', 'Debit': '0.00', 'Credit': '15.50', 'Balance': '9264.75'}
        ]
        
        exporter = StatementExporter()
        
        if format_type.lower() == "excel":
            # Generate Excel
            excel_data = exporter.to_excel(
                transactions=sample_transactions,
                bank_name="Sample Bank",
                statement_period="June 2025"
            )
            
            filename = "Sample_Bank_Statement_June_2025.xlsx"
            
            def generate():
                yield excel_data
            
            return StreamingResponse(
                generate(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        elif format_type.lower() == "csv":
            # Generate CSV
            csv_data = exporter.to_csv(transactions=sample_transactions)
            
            filename = "Sample_Bank_Statement_June_2025.csv"
            
            def generate():
                yield csv_data.encode('utf-8')
            
            return StreamingResponse(
                generate(),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        else:
            raise HTTPException(status_code=400, detail="Invalid format_type. Use 'excel' or 'csv'")
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating sample export: {str(e)}")

@app.post("/bank-specific-extract")
async def bank_specific_extract(file: UploadFile = File(...)):
    """Template-based bank statement parser using regex patterns from Claude guide"""
    try:
        print(f"[Template Extract] Starting template-based extraction for file: {file.filename}")
        
        pdf_bytes = await file.read()
        all_text = ""
        
        # Extract text from PDF
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            total_pages = len(pdf.pages)
            print(f"[Template Extract] Processing {total_pages} pages")
            
            for page_num, pdf_page in enumerate(pdf.pages):
                print(f"[Template Extract] Processing page {page_num + 1}/{total_pages}")
                
                if is_digital_pdf(pdf_page):
                    print(f"[Template Extract] Page {page_num + 1} is digital, using pdfplumber")
                    text = pdf_page.extract_text()
                    if text:
                        all_text += text + "\n"
                        print(f"[Template Extract] Extracted text from digital PDF page.")
                else:
                    print(f"[Template Extract] Page {page_num + 1} is scanned, skipping for now")
        
        print(f"[Template Extract] Total text length: {len(all_text)} characters")
        
        # Debug: Print first 500 characters of extracted text
        print(f"[Template Extract] DEBUG - First 500 chars of text:")
        print(repr(all_text[:500]))
        
        if not all_text.strip():
            return {"success": False, "error": "No text extracted from PDF", "transactions": []}
        
        # Phase 1: Bank Detection using Claude guide patterns
        def detect_bank(text):
            bank_patterns = {
                'HDFC': r'HDFC BANK LIMITED',
                'ICICI': r'ICICI BANK',
                'IDFC': r'IDFC FIRST BANK'
            }
            for bank, pattern in bank_patterns.items():
                if re.search(pattern, text, re.IGNORECASE):
                    return bank
            return None
        
        detected_bank = detect_bank(all_text)
        print(f"[Template Extract] Detected bank: {detected_bank}")
        
        if not detected_bank:
            return {"success": False, "error": "Could not detect bank type", "transactions": []}
        
        # Phase 2: Template-based parsing with Claude guide patterns
        templates = {
            'HDFC': {
                'date_pattern': r'(\d{2}/\d{2}/\d{4})',
                'amount_pattern': r'(\d+,?\d*\.\d{2})',
                'transaction_start': r'Date\s+Narration',
                'transaction_pattern': r'(\d{2}/\d{2}/\d{4})\s+(.+?)\s+(\w+)\s+(\d{2}/\d{2}/\d{4})\s+(\d+\.\d{2})\s+(\d+\.\d{2})\s+(\d+\.\d{2})'
            },
            'ICICI': {
                'transaction_pattern': r'(\d{2}-\d{2}-\d{4})\s+(\w+)\s+(.+?)\s+(\d+,?\d*\.\d{2})?\s+(\d+,?\d*\.\d{2})?\s+(\d+,?\d*\.\d{2})'
            },
            'IDFC': {
                'transaction_pattern': r'(\d{2}/\d{2}/\d{4})\s+(.+?)\s+(\d+,?\d*\.\d{2})\s+(\d+,?\d*\.\d{2})\s+(\d+,?\d*\.\d{2})'
            }
        }
        
        # Phase 3: Extract transactions using bank-specific patterns
        def extract_transactions(text, bank_type):
            transactions = []
            if bank_type not in templates:
                return transactions
                
            pattern = templates[bank_type]['transaction_pattern']
            matches = re.findall(pattern, text, re.MULTILINE)
            
            print(f"[Template Extract] Found {len(matches)} transaction matches for {bank_type}")
            
            for i, match in enumerate(matches):
                try:
                    if bank_type == 'HDFC':
                        transaction = {
                            'Date': match[0],
                            'Narration': match[1],
                            'Chq_Ref_No': match[2],
                            'Value_Date': match[3],
                            'Withdrawal_Amount': match[4],
                            'Deposit_Amount': match[5],
                            'Closing_Balance': match[6]
                        }
                    elif bank_type == 'ICICI':
                        transaction = {
                            'Date': match[0],
                            'Mode': match[1],
                            'Particulars': match[2],
                            'Deposits': match[3] if match[3] else '0.00',
                            'Withdrawals': match[4] if match[4] else '0.00',
                            'Balance': match[5]
                        }
                    elif bank_type == 'IDFC':
                        transaction = {
                            'Date_and_Time': match[0],
                            'Transaction_Details': match[1],
                            'Withdrawals_INR': match[2],
                            'Deposits_INR': match[3],
                            'Balance_INR': match[4]
                        }
                    
                    transactions.append(transaction)
                    print(f"[Template Extract] Transaction {i+1}: {transaction}")
                    
                except Exception as e:
                    print(f"[Template Extract] Error parsing transaction {i+1}: {e}")
                    continue
            
            return transactions
        
        transactions = extract_transactions(all_text, detected_bank)
        
        # Phase 4: Preserve original headers (no standardization)
        def preserve_original_output(transactions, bank_type):
            """Return transactions with original bank-specific headers preserved"""
            preserved = []
            for txn in transactions:
                try:
                    # Keep the original field names exactly as they appear in the PDF
                    preserved.append(txn)
                except Exception as e:
                    print(f"[Template Extract] Error preserving transaction: {e}")
                    continue
            
            return preserved
        
        preserved_transactions = preserve_original_output(transactions, detected_bank)
        
        print(f"[Template Extract] Successfully extracted {len(preserved_transactions)} transactions from {detected_bank} statement")
        
        return {
            "success": True,
            "bank_type": detected_bank,
            "transactions": preserved_transactions,
            "total_transactions": len(preserved_transactions)
        }
        
    except Exception as e:
        print(f"[Template Extract] Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e), "transactions": []}

@app.post("/intelligent-bank-extract")
async def intelligent_bank_extract(file: UploadFile = File(...)):
    """
    V2025.07.21.01: Intelligent Universal Bank Statement Parser
    
    Features:
    - Multi-page transaction extraction (ALL transactions from ALL pages)
    - Adaptive header mapping (preserves original headers exactly)
    - Zero truncation (complete field values preserved)
    - Works with any bank statement format (ICICI, IDFC, HDFC, etc.)
    - Smart transaction detection using AI patterns
    """
    try:
        from .intelligent_parser import parse_bank_statement_intelligent
        
        print(f"[Intelligent Parser V2025.07.21.01] Processing: {file.filename}")
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_file_path = tmp_file.name
        
        # Use intelligent parser
        results = parse_bank_statement_intelligent(tmp_file_path)
        
        # Clean up temp file
        os.unlink(tmp_file_path)
        
        print(f"[Intelligent Parser] Success: {results['total_transactions']} transactions from {results['total_pages']} pages")
        
        return results
        
    except Exception as e:
        print(f"[Intelligent Parser] Error: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Clean up temp file on error
        if 'tmp_file_path' in locals():
            try:
                os.unlink(tmp_file_path)
            except:
                pass
                
        return {
            "success": False, 
            "error": str(e), 
            "parser_version": "V2025.07.21.01_Intelligent"
        }

# Enhanced Transaction Extraction Endpoints

@app.post("/detect-transaction-pages")
async def detect_transaction_pages(
    file: UploadFile = File(...),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Detect transaction pages using enhanced parser v2.0
    Returns first transaction page with pattern analysis
    """
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400, 
            detail="Only PDF files are allowed"
        )
    
    try:
        # Save uploaded file temporarily
        temp_filename = f"detect_{uuid.uuid4()}.pdf"
        temp_filepath = os.path.join(tempfile.gettempdir(), temp_filename)
        
        with open(temp_filepath, "wb") as temp_file:
            content = await file.read()
            temp_file.write(content)
        
        # Use enhanced detector
        detector = EnhancedTransactionDetector()
        detection_result = detector.detect_first_transaction_page(temp_filepath)
        
        # Clean up temp file
        os.unlink(temp_filepath)
        
        return {
            "filename": file.filename,
            "detection_result": detection_result,
            "message": "Transaction page detection complete"
        }
        
    except Exception as e:
        # Clean up temp file if it exists
        if 'temp_filepath' in locals() and os.path.exists(temp_filepath):
            os.unlink(temp_filepath)
        
        raise HTTPException(
            status_code=500,
            detail=f"Error detecting transaction pages: {str(e)}"
        )

@app.post("/extract-complete-transactions")
async def extract_complete_transactions(
    file: UploadFile = File(...),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Complete transaction extraction workflow using pattern rule applicator
    Detects structure, generates pattern rules, and extracts from all pages
    """
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400, 
            detail="Only PDF files are allowed"
        )
    
    try:
        # Save uploaded file temporarily
        temp_filename = f"extract_{uuid.uuid4()}.pdf"
        temp_filepath = os.path.join(tempfile.gettempdir(), temp_filename)
        
        with open(temp_filepath, "wb") as temp_file:
            content = await file.read()
            temp_file.write(content)
        
        # Use pattern rule applicator for complete workflow
        applicator = PatternRuleApplicator()
        extraction_result = applicator.generate_and_apply_pattern_rule(temp_filepath)
        
        # Clean up temp file
        os.unlink(temp_filepath)
        
        # Convert ExtractionResult to JSON-serializable format
        result_dict = {
            "filename": file.filename,
            "success": extraction_result.success,
            "total_transactions": extraction_result.total_transactions,
            "transactions": extraction_result.transactions,
            "pages_processed": extraction_result.pages_processed,
            "pattern_rule_used": extraction_result.pattern_rule_used,
            "extraction_summary": extraction_result.extraction_summary,
            "errors": extraction_result.errors,
            "message": f"Extraction complete: {extraction_result.total_transactions} transactions from {len(extraction_result.pages_processed)} pages"
        }
        
        return result_dict
        
    except Exception as e:
        # Clean up temp file if it exists
        if 'temp_filepath' in locals() and os.path.exists(temp_filepath):
            os.unlink(temp_filepath)
        
        raise HTTPException(
            status_code=500,
            detail=f"Error extracting transactions: {str(e)}"
        )

@app.post("/apply-custom-pattern")
async def apply_custom_pattern(
    file: UploadFile = File(...),
    pattern_name: str = Form(...),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Apply a saved custom pattern rule to extract transactions
    """
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400, 
            detail="Only PDF files are allowed"
        )
    
    try:
        # Load the saved pattern
        pattern_manager = EnhancedPatternRuleManager()
        pattern_rule = pattern_manager.get_pattern(pattern_name)
        
        if not pattern_rule:
            raise HTTPException(
                status_code=404,
                detail=f"Pattern '{pattern_name}' not found"
            )
        
        # Save uploaded file temporarily
        temp_filename = f"custom_{uuid.uuid4()}.pdf"
        temp_filepath = os.path.join(tempfile.gettempdir(), temp_filename)
        
        with open(temp_filepath, "wb") as temp_file:
            content = await file.read()
            temp_file.write(content)
        
        # Apply the custom pattern (simplified version)
        applicator = PatternRuleApplicator()
        
        # First detect to get basic structure
        detector = EnhancedTransactionDetector()
        detection_result = detector.detect_first_transaction_page(temp_filepath)
        
        if not detection_result['success']:
            raise HTTPException(
                status_code=400,
                detail="Could not detect transaction pages in the PDF"
            )
        
        # Apply the custom pattern to all pages
        extraction_result = applicator._apply_pattern_rule_to_all_pages(
            temp_filepath, pattern_rule, detection_result
        )
        
        # Clean up temp file
        os.unlink(temp_filepath)
        
        # Convert to JSON-serializable format
        result_dict = {
            "filename": file.filename,
            "pattern_name": pattern_name,
            "success": extraction_result.success,
            "total_transactions": extraction_result.total_transactions,
            "transactions": extraction_result.transactions,
            "pages_processed": extraction_result.pages_processed,
            "pattern_rule_used": extraction_result.pattern_rule_used,
            "extraction_summary": extraction_result.extraction_summary,
            "errors": extraction_result.errors,
            "message": f"Custom pattern applied: {extraction_result.total_transactions} transactions from {len(extraction_result.pages_processed)} pages"
        }
        
        return result_dict
        
    except Exception as e:
        # Clean up temp file if it exists
        if 'temp_filepath' in locals() and os.path.exists(temp_filepath):
            os.unlink(temp_filepath)
        
        raise HTTPException(
            status_code=500,
            detail=f"Error applying custom pattern: {str(e)}"
        )

@app.get("/saved-patterns")
async def get_saved_patterns(current_user: UserModel = Depends(get_current_user)):
    """
    Get list of all saved pattern rules
    """
    try:
        pattern_manager = EnhancedPatternRuleManager()
        patterns = pattern_manager.list_patterns()
        
        return {
            "patterns": patterns,
            "total_patterns": len(patterns),
            "message": f"Found {len(patterns)} saved patterns"
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error retrieving saved patterns: {str(e)}"
        )

@app.post("/save-pattern")
async def save_pattern_rule(
    pattern_name: str = Form(...),
    headers: str = Form(...),  # JSON string of headers list
    layout_mode: str = Form("text-aligned"),
    format_type: str = Form("text_based"),
    bank_name: str = Form(""),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Save a custom pattern rule for reuse
    """
    try:
        # Parse headers from JSON string
        import json
        headers_list = json.loads(headers)
        
        if not isinstance(headers_list, list) or not headers_list:
            raise HTTPException(
                status_code=400,
                detail="Headers must be a non-empty list"
            )
        
        # Create pattern rule
        pattern_rule = PatternRule(
            column_count=len(headers_list),
            header_keywords=headers_list,
            row_gap_tolerance=10.0,
            font_size_range=(8.0, 14.0),
            first_column_pattern=r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}',
            layout_mode=layout_mode,
            header_positions=[(i * 120, (i + 1) * 120) for i in range(len(headers_list))],
            row_height=18.0,
            format_type=format_type
        )
        
        # Save pattern
        pattern_manager = EnhancedPatternRuleManager()
        success = pattern_manager.save_pattern(pattern_name, pattern_rule, bank_name)
        
        if success:
            return {
                "success": True,
                "pattern_name": pattern_name,
                "headers": headers_list,
                "bank_name": bank_name,
                "message": f"Pattern '{pattern_name}' saved successfully"
            }
        else:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to save pattern '{pattern_name}'"
            )
        
    except json.JSONDecodeError:
        raise HTTPException(
            status_code=400,
            detail="Invalid JSON format for headers"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error saving pattern: {str(e)}"
        )

class ExtractionStatisticsRequest(BaseModel):
    """Request model for extraction statistics"""
    success: bool
    total_transactions: int
    transactions: List[Dict[str, Any]]
    pages_processed: List[int]
    pattern_rule_used: Dict[str, Any]
    extraction_summary: Dict[str, Any]
    errors: List[str]

@app.post("/extraction-statistics")
async def get_extraction_statistics(
    extraction_data: ExtractionStatisticsRequest,
    current_user: UserModel = Depends(get_current_user)
):
    """
    Generate detailed statistics from extraction results
    """
    try:
        # Convert request to ExtractionResult object
        extraction_result = ExtractionResult(
            success=extraction_data.success,
            total_transactions=extraction_data.total_transactions,
            transactions=extraction_data.transactions,
            pages_processed=extraction_data.pages_processed,
            pattern_rule_used=extraction_data.pattern_rule_used,
            extraction_summary=extraction_data.extraction_summary,
            errors=extraction_data.errors
        )
        
        # Generate statistics
        applicator = PatternRuleApplicator()
        statistics = applicator.get_extraction_statistics(extraction_result)
        
        return {
            "statistics": statistics,
            "message": "Statistics generated successfully"
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error generating statistics: {str(e)}"
        )

# Smart PDF Parser specific endpoints (isolated from other features)
def smart_extract_text_blocks(pdf_bytes):
    """
    Extract meaningful text blocks by grouping words into rows and phrases
    ONLY for Smart PDF Parser - does not affect other features
    """
    results = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages):
            # Extract words with positioning
            words = page.extract_words(x_tolerance=1, y_tolerance=1)
            
            if not words:
                results.append({"page": page_num, "blocks": []})
                continue
            
            # Group words into meaningful text blocks
            text_blocks = smart_group_words_into_blocks(words, page_num)
            
            results.append({
                "page": page_num,
                "blocks": text_blocks
            })
    return results

def smart_group_words_into_blocks(words, page_num):
    """
    Group words into meaningful text blocks using layout consistency rules
    """
    if not words:
        return []
    
    # Step 1: Snap Y-coordinates into row buckets
    rows = smart_group_words_by_rows(words, y_tolerance=3)
    
    # Step 2: For each row, group nearby words into phrases
    text_blocks = []
    for row_words in rows:
        phrases = smart_group_words_into_phrases(row_words, x_tolerance=15)
        text_blocks.extend(phrases)
    
    # Step 3: Filter out noise (single chars, punctuation, etc.)
    filtered_blocks = smart_filter_meaningful_blocks(text_blocks)
    
    return filtered_blocks

def smart_group_words_by_rows(words, y_tolerance=3):
    """
    Group words that are on the same horizontal line (row)
    """
    rows = {}
    
    for word in words:
        y_center = (word["top"] + word["bottom"]) / 2
        
        # Find existing row within tolerance
        matched_row = None
        for existing_y in rows.keys():
            if abs(y_center - existing_y) <= y_tolerance:
                matched_row = existing_y
                break
        
        if matched_row is not None:
            rows[matched_row].append(word)
        else:
            rows[y_center] = [word]
    
    # Sort rows top to bottom and words left to right within each row
    sorted_rows = []
    for y in sorted(rows.keys()):
        row_words = sorted(rows[y], key=lambda w: w["x0"])
        sorted_rows.append(row_words)
    
    return sorted_rows

def smart_group_words_into_phrases(row_words, x_tolerance=20):
    """
    Group words in a row into phrases based on horizontal proximity
    """
    if not row_words:
        return []
    
    phrases = []
    current_phrase = [row_words[0]]
    
    for i in range(1, len(row_words)):
        prev_word = row_words[i-1]
        curr_word = row_words[i]
        
        # Check if current word is close enough to be part of the same phrase
        gap = curr_word["x0"] - prev_word["x1"]
        
        if gap <= x_tolerance:
            current_phrase.append(curr_word)
        else:
            # Finish current phrase and start new one
            if current_phrase:
                phrases.append(smart_create_phrase_block(current_phrase))
            current_phrase = [curr_word]
    
    # Add the last phrase
    if current_phrase:
        phrases.append(smart_create_phrase_block(current_phrase))
    
    return phrases

def smart_create_phrase_block(words):
    """
    Create a single text block from a group of words
    """
    if not words:
        return None
    
    # Combine text
    text_parts = [word["text"] for word in words]
    combined_text = " ".join(text_parts)
    
    # Calculate bounding box
    x0 = min(word["x0"] for word in words)
    y0 = min(word["top"] for word in words)  
    x1 = max(word["x1"] for word in words)
    y1 = max(word["bottom"] for word in words)
    
    block = {
        "text": combined_text,
        "x0": x0,
        "y0": y0,
        "x1": x1,
        "y1": y1
    }
    
    return block

def smart_filter_meaningful_blocks(blocks):
    """
    Filter out noise blocks (single chars, punctuation, etc.)
    """
    filtered = []
    
    for block in blocks:
        if not block or not block.get("text"):
            continue
            
        text = block["text"].strip()
        
        # Skip empty blocks
        if len(text) < 1:
            continue
            
        # Skip blocks with only punctuation or single characters (unless meaningful numbers/letters)
        if len(text) == 1 and not text.isalnum():
            continue
            
        # Skip blocks that are just whitespace or special chars
        if not any(c.isalnum() for c in text):
            continue
            
        # Skip extremely small blocks (likely noise) - very permissive
        width = block["x1"] - block["x0"]
        height = block["y1"] - block["y0"]
        # Allow zero width (stacked text) but require some height
        if height < 1:
            continue
            
        # Allow single meaningful characters (letters, numbers)
        # Skip only if it's a single non-alphanumeric character
        if len(text.strip()) == 1 and not text.isalnum():
            continue
        
        filtered.append(block)
    
    return filtered

@app.post("/smart-extract-blocks")
async def smart_extract_blocks(file: UploadFile = File(...)):
    """
    Extract text blocks for Smart PDF Parser with improved grouping and filtering.
    This endpoint is ONLY for Smart PDF Parser and does not affect other features.
    """
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF.")
    
    pdf_bytes = await file.read()
    try:
        # Use the new smart extraction function
        blocks = smart_extract_text_blocks(pdf_bytes)
        return JSONResponse(content={"pages": blocks})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
