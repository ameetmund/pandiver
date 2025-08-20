import os
import tempfile
import uuid
from typing import List, Dict, Any, Optional
from datetime import datetime
from fastapi import APIRouter, File, UploadFile, HTTPException, Depends, Form, Body
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from celery.result import AsyncResult

from ..auth import get_current_user
from ..models import User as UserModel
from .models import (
    SingleFileProcessRequest, BulkFileProcessRequest, JobResponse, 
    JobStatusResponse, BulkJobResponse, BulkJobStatusResponse,
    ExtractionResult, ProcessingStatus, ExtractionMethod
)
from .tasks import process_single_file_api, process_bulk_files_api

router = APIRouter()
security = HTTPBearer()


@router.post("/process", response_model=JobResponse)
async def process_single_file(
    file: UploadFile = File(...),
    extraction_method: ExtractionMethod = Form(ExtractionMethod.AI_BANK_PARSER),
    webhook_url: Optional[str] = Form(None),
    webhook_events: Optional[str] = Form(None),  # JSON string
    metadata: Optional[str] = Form("{}"),  # JSON string
    current_user: UserModel = Depends(get_current_user)
):
    """
    Process a single PDF file for transaction extraction.
    
    **Parameters:**
    - `file`: PDF file to process (required)
    - `extraction_method`: Method to use ('smart', 'intelligent', 'textract', 'manual') 
    - `webhook_url`: Optional URL to receive webhooks on job completion
    - `webhook_events`: JSON array of events to webhook (e.g., ["processing.completed", "processing.failed"])
    - `metadata`: JSON object with additional metadata
    
    **Returns:**
    - Job ID for tracking the processing status
    - Use `/api/v1/jobs/{job_id}/status` to check progress
    
    **Example curl:**
    ```bash
    curl -X POST "http://localhost:8000/api/v1/process" \
         -H "Authorization: Bearer YOUR_JWT_TOKEN" \
         -F "file=@bank_statement.pdf" \
         -F "extraction_method=smart" \
         -F "webhook_url=https://yourdomain.com/webhooks/pandiver"
    ```
    """
    # Validate file
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400, 
            detail="Only PDF files are supported"
        )
    
    try:
        # Parse optional JSON parameters
        import json
        webhook_events_list = None
        if webhook_events:
            webhook_events_list = json.loads(webhook_events)
        
        metadata_dict = json.loads(metadata) if metadata else {}
        
        # Save uploaded file temporarily
        file_extension = ".pdf"
        temp_filename = f"api_process_{uuid.uuid4()}{file_extension}"
        temp_filepath = os.path.join(tempfile.gettempdir(), temp_filename)
        
        with open(temp_filepath, "wb") as temp_file:
            content = await file.read()
            temp_file.write(content)
        
        # Queue processing task
        task = process_single_file_api.delay(
            temp_filepath,
            extraction_method.value,
            webhook_url,
            webhook_events_list,
            metadata_dict
        )
        
        return JobResponse(
            job_id=task.id,
            status=ProcessingStatus.PENDING,
            message="File uploaded successfully. Processing started.",
            created_at=datetime.utcnow(),
            webhook_url=webhook_url,
            metadata=metadata_dict
        )
        
    except json.JSONDecodeError as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid JSON in parameters: {str(e)}"
        )
    except Exception as e:
        # Clean up temp file
        if 'temp_filepath' in locals() and os.path.exists(temp_filepath):
            os.unlink(temp_filepath)
        
        raise HTTPException(
            status_code=500,
            detail=f"Error starting processing job: {str(e)}"
        )


@router.post("/process-bulk", response_model=BulkJobResponse)
async def process_bulk_files(
    files: List[UploadFile] = File(...),
    extraction_method: ExtractionMethod = Form(ExtractionMethod.AI_BANK_PARSER),
    webhook_url: Optional[str] = Form(None),
    webhook_events: Optional[str] = Form(None),  # JSON string
    metadata: Optional[str] = Form("{}"),  # JSON string
    max_concurrent_jobs: int = Form(3),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Process multiple PDF files concurrently.
    
    **Parameters:**
    - `files`: Multiple PDF files to process (required)
    - `extraction_method`: Method to use for all files
    - `webhook_url`: Optional URL to receive webhooks on bulk job completion
    - `webhook_events`: JSON array of events to webhook
    - `metadata`: JSON object with additional metadata
    - `max_concurrent_jobs`: Maximum concurrent processing jobs (1-10)
    
    **Returns:**
    - Bulk job ID and individual job IDs for tracking
    
    **Example curl:**
    ```bash
    curl -X POST "http://localhost:8000/api/v1/process-bulk" \
         -H "Authorization: Bearer YOUR_JWT_TOKEN" \
         -F "files=@statement1.pdf" \
         -F "files=@statement2.pdf" \
         -F "files=@statement3.pdf" \
         -F "extraction_method=smart" \
         -F "max_concurrent_jobs=2"
    ```
    """
    # Validate files
    if not files or len(files) == 0:
        raise HTTPException(
            status_code=400,
            detail="At least one file must be provided"
        )
    
    if len(files) > 50:  # Reasonable limit
        raise HTTPException(
            status_code=400,
            detail="Maximum 50 files allowed per bulk request"
        )
    
    for file in files:
        if not file.filename or not file.filename.lower().endswith('.pdf'):
            raise HTTPException(
                status_code=400, 
                detail=f"File '{file.filename}' is not a PDF"
            )
    
    try:
        # Parse optional JSON parameters
        import json
        webhook_events_list = None
        if webhook_events:
            webhook_events_list = json.loads(webhook_events)
        
        metadata_dict = json.loads(metadata) if metadata else {}
        
        # Save all uploaded files temporarily
        temp_filepaths = []
        
        for file in files:
            file_extension = ".pdf"
            temp_filename = f"api_bulk_{uuid.uuid4()}{file_extension}"
            temp_filepath = os.path.join(tempfile.gettempdir(), temp_filename)
            
            with open(temp_filepath, "wb") as temp_file:
                content = await file.read()
                temp_file.write(content)
            
            temp_filepaths.append(temp_filepath)
        
        # Queue bulk processing task
        bulk_task = process_bulk_files_api.delay(
            temp_filepaths,
            extraction_method.value,
            webhook_url,
            webhook_events_list,
            metadata_dict,
            max_concurrent_jobs
        )
        
        return BulkJobResponse(
            bulk_job_id=bulk_task.id,
            individual_job_ids=[],  # Individual job IDs will be available in status
            total_files=len(files),
            status=ProcessingStatus.PENDING,
            created_at=datetime.utcnow(),
            webhook_url=webhook_url
        )
        
    except json.JSONDecodeError as e:
        # Clean up temp files
        for temp_filepath in temp_filepaths:
            if os.path.exists(temp_filepath):
                os.unlink(temp_filepath)
        
        raise HTTPException(
            status_code=400,
            detail=f"Invalid JSON in parameters: {str(e)}"
        )
    except Exception as e:
        # Clean up temp files
        if 'temp_filepaths' in locals():
            for temp_filepath in temp_filepaths:
                if os.path.exists(temp_filepath):
                    os.unlink(temp_filepath)
        
        raise HTTPException(
            status_code=500,
            detail=f"Error starting bulk processing job: {str(e)}"
        )


@router.get("/jobs/{job_id}/status", response_model=JobStatusResponse)
async def get_job_status(
    job_id: str,
    current_user: UserModel = Depends(get_current_user)
):
    """
    Get the status of a processing job.
    
    **Parameters:**
    - `job_id`: The job ID returned from the process endpoint
    
    **Returns:**
    - Current job status, progress, and results (if completed)
    
    **Example curl:**
    ```bash
    curl -H "Authorization: Bearer YOUR_JWT_TOKEN" \
         "http://localhost:8000/api/v1/jobs/abc123/status"
    ```
    """
    try:
        task_result = AsyncResult(job_id)
        
        # Determine status and progress
        if task_result.state == 'PENDING':
            status = ProcessingStatus.PENDING
            progress = 0
            current_step = "Waiting to start"
            result_data = None
            error_message = None
        elif task_result.state == 'IN_PROGRESS' or task_result.state == 'PROGRESS':
            status = ProcessingStatus.IN_PROGRESS
            info = task_result.info or {}
            progress = info.get('progress', 50)
            current_step = info.get('current_step', 'Processing...')
            result_data = None
            error_message = None
        elif task_result.state == 'SUCCESS':
            status = ProcessingStatus.COMPLETED
            progress = 100
            current_step = "Completed"
            result_data = task_result.result
            error_message = None
        elif task_result.state == 'FAILURE':
            status = ProcessingStatus.FAILED
            progress = 0
            current_step = "Failed"
            result_data = None
            error_message = str(task_result.info) if task_result.info else "Unknown error"
        else:
            status = ProcessingStatus.PENDING
            progress = 0
            current_step = f"Status: {task_result.state}"
            result_data = None
            error_message = None
        
        # Get task creation time (approximation)
        created_at = datetime.utcnow()  # This would ideally come from job storage
        
        return JobStatusResponse(
            job_id=job_id,
            status=status,
            progress_percentage=progress,
            current_step=current_step,
            result=result_data,
            error_message=error_message,
            created_at=created_at,
            updated_at=datetime.utcnow(),
            estimated_completion=None  # Could be calculated based on progress
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error retrieving job status: {str(e)}"
        )


@router.get("/jobs/{job_id}/result")
async def get_job_result(
    job_id: str,
    format: str = "json",  # json, csv, xlsx
    current_user: UserModel = Depends(get_current_user)
):
    """
    Get the result of a completed processing job.
    
    **Parameters:**
    - `job_id`: The job ID
    - `format`: Result format ('json', 'csv', 'xlsx', 'txt') - default: 'json'
    
    **Returns:**
    - Job results in the specified format
    
    **Example curl:**
    ```bash
    curl -H "Authorization: Bearer YOUR_JWT_TOKEN" \
         "http://localhost:8000/api/v1/jobs/abc123/result?format=csv"
    ```
    """
    try:
        task_result = AsyncResult(job_id)
        
        if task_result.state != 'SUCCESS':
            raise HTTPException(
                status_code=400,
                detail=f"Job is not completed yet. Current status: {task_result.state}"
            )
        
        result_data = task_result.result
        
        if format.lower() == "json":
            return result_data
        elif format.lower() == "csv":
            # Convert to CSV format
            return export_to_csv(result_data)
        elif format.lower() == "xlsx":
            # Convert to Excel format - returns ZIP like CSV
            return export_to_excel(result_data)
        elif format.lower() == "txt":
            # Convert to TXT format
            return export_to_txt(result_data)
        else:
            raise HTTPException(
                status_code=400,
                detail="Invalid format. Supported formats: json, csv, xlsx, txt"
            )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error retrieving job result: {str(e)}"
        )


@router.delete("/jobs/{job_id}")
async def cancel_job(
    job_id: str,
    current_user: UserModel = Depends(get_current_user)
):
    """
    Cancel a pending or in-progress job.
    
    **Parameters:**
    - `job_id`: The job ID to cancel
    
    **Returns:**
    - Confirmation of cancellation
    
    **Example curl:**
    ```bash
    curl -X DELETE -H "Authorization: Bearer YOUR_JWT_TOKEN" \
         "http://localhost:8000/api/v1/jobs/abc123"
    ```
    """
    try:
        from celery import current_app
        current_app.control.revoke(job_id, terminate=True)
        
        return {
            "job_id": job_id,
            "status": "cancelled",
            "message": "Job has been cancelled"
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error cancelling job: {str(e)}"
        )


@router.get("/jobs", response_model=List[JobStatusResponse])
async def list_user_jobs(
    limit: int = 10,
    offset: int = 0,
    status: Optional[ProcessingStatus] = None,
    current_user: UserModel = Depends(get_current_user)
):
    """
    List processing jobs for the current user.
    
    **Parameters:**
    - `limit`: Maximum number of jobs to return (default: 10, max: 100)
    - `offset`: Number of jobs to skip (default: 0)
    - `status`: Filter by job status (optional)
    
    **Returns:**
    - List of user's jobs with their current status
    
    **Note:** This is a simplified implementation. In production, you'd want to 
    store job metadata in a database with user association.
    """
    # This is a placeholder implementation
    # In a real system, you'd query a database of jobs associated with the user
    
    if limit > 100:
        limit = 100
    
    # For now, return empty list with explanation
    return []


@router.post("/test-export")
async def test_export_functions():
    """
    Test endpoint to verify export functions work correctly - TEMPORARY
    """
    # Mock data that looks like what the UI produces
    mock_result_data = {
        'success': True,
        'tables': [
            {
                'table_id': 1,
                'page_number': 0,
                'headers': ['Date', 'Description', 'Debit', 'Credit', 'Balance'],
                'rows': [
                    ['2024-01-01', 'Opening Balance', '', '1000.00', '1000.00'],
                    ['2024-01-02', 'ATM Withdrawal', '200.00', '', '800.00'],
                    ['2024-01-03', 'Deposit', '', '500.00', '1300.00']
                ]
            },
            {
                'table_id': 2,
                'page_number': 1,
                'headers': ['Date', 'Transaction', 'Amount'],
                'rows': [
                    ['2024-01-04', 'Transfer', '250.00'],
                    ['2024-01-05', 'Payment', '150.00']
                ]
            }
        ],
        'total_tables': 2
    }
    
    try:
        # Test CSV export
        csv_result = export_to_csv(mock_result_data)
        
        # Test Excel export
        excel_result = export_to_excel(mock_result_data)
        
        return {
            'message': 'Export functions tested successfully',
            'csv_type': str(type(csv_result)),
            'excel_type': str(type(excel_result)),
            'csv_size': len(csv_result.body) if hasattr(csv_result, 'body') else 'N/A',
            'excel_size': len(excel_result) if isinstance(excel_result, bytes) else 'N/A',
            'mock_data_structure': list(mock_result_data.keys())
        }
        
    except Exception as e:
        return {
            'error': str(e),
            'error_type': str(type(e)),
            'message': 'Export functions failed'
        }


@router.post("/test-bulk-export")
async def test_bulk_export_functions():
    """
    Test endpoint to verify bulk export functions work correctly - TEMPORARY
    """
    # Mock bulk data structure like what bulk jobs return
    mock_bulk_data = {
        'bulk_job_id': 'test-bulk-123',
        'status': 'completed',
        'total_files': 2,
        'completed_files': 2,
        'failed_files': 0,
        'individual_results': [
            {
                'job_id': 'file1-job',
                'status': 'completed',
                'file_name': 'file1.pdf',
                'tables': [
                    {
                        'table_id': 1,
                        'headers': ['Date', 'Amount', 'Description'],
                        'rows': [
                            ['2024-01-01', '100.00', 'Payment 1'],
                            ['2024-01-02', '200.00', 'Payment 2']
                        ]
                    }
                ],
                'total_tables': 1
            },
            {
                'job_id': 'file2-job', 
                'status': 'completed',
                'file_name': 'file2.pdf',
                'tables': [
                    {
                        'table_id': 1,
                        'headers': ['Date', 'Amount', 'Description'],
                        'rows': [
                            ['2024-01-03', '300.00', 'Payment 3'],
                            ['2024-01-04', '400.00', 'Payment 4']
                        ]
                    }
                ],
                'total_tables': 1
            }
        ]
    }
    
    try:
        # Test TXT export with bulk data
        txt_result = export_to_txt(mock_bulk_data)
        
        return {
            'message': 'Bulk export functions tested successfully',
            'txt_type': str(type(txt_result)),
            'mock_data_structure': list(mock_bulk_data.keys()),
            'individual_results_count': len(mock_bulk_data['individual_results']),
            'total_individual_tables': sum(len(result.get('tables', [])) for result in mock_bulk_data['individual_results'])
        }
        
    except Exception as e:
        return {
            'error': str(e),
            'error_type': str(type(e)),
            'message': 'Bulk export functions failed'
        }


# Helper functions for format conversion
def export_to_csv(result_data: Dict[str, Any]):
    """Convert job result to CSV format - EXACT same as UI behavior (returns ZIP)"""
    from ..textract_endpoints import generate_csv_content
    from fastapi.responses import Response
    import zipfile
    import io
    
    # Handle UI-compatible structure
    tables_to_export = []
    
    if 'tables' in result_data and result_data.get('tables'):
        # AI Bank Parser result - same as UI
        tables_to_export = result_data.get('tables', [])
    elif 'forms_data' in result_data and result_data.get('forms_data'):
        # Form Data Parser result - convert forms to tables
        forms_data = result_data.get('forms_data', [])
        for page_data in forms_data:
            headers = page_data.get('headers', [])
            rows = page_data.get('data', [])
            if headers and rows:
                tables_to_export.append({
                    'table_id': len(tables_to_export) + 1,
                    'headers': headers,
                    'rows': rows,
                    'page_number': page_data.get('page_number', 0)
                })
    elif 'individual_results' in result_data:
        # Bulk job result - combine all tables from individual results
        for individual_result in result_data.get('individual_results', []):
            if 'tables' in individual_result:
                tables_to_export.extend(individual_result['tables'])
            elif 'forms_data' in individual_result:
                forms_data = individual_result['forms_data']
                for page_data in forms_data:
                    headers = page_data.get('headers', [])
                    rows = page_data.get('data', [])
                    if headers and rows:
                        tables_to_export.append({
                            'table_id': len(tables_to_export) + 1,
                            'headers': headers,
                            'rows': rows,
                            'page_number': page_data.get('page_number', 0)
                        })
    
    if not tables_to_export:
        return Response(
            content=f"No tables found for export. Keys: {list(result_data.keys())}",
            media_type='text/plain'
        )
    
    # Create ZIP with multiple CSV files (same as UI)
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for table_idx, table in enumerate(tables_to_export):
            table_id = table.get('table_id', table_idx + 1)
            page_number = table.get('page_number', 0) + 1  # Convert to 1-based
            headers = table.get('headers', [])
            rows = table.get('rows', [])
            
            if not headers or not rows:
                continue
                
            filename = f"table_{table_id}_page_{page_number}.csv"
            content = generate_csv_content(headers, rows)
            zip_file.writestr(filename, content)
    
    zip_buffer.seek(0)
    
    return Response(
        content=zip_buffer.getvalue(),
        media_type='application/zip',
        headers={'Content-Disposition': 'attachment; filename=bank-statement-tables.zip'}
    )


def export_to_excel(result_data: Dict[str, Any]):
    """Convert job result to Excel format - same as UI behavior (returns ZIP)"""
    from ..textract_endpoints import generate_excel_content
    import io
    import zipfile
    import openpyxl
    from datetime import datetime
    
    # Handle different result structures (same logic as CSV)
    tables_to_export = []
    
    if 'tables' in result_data and result_data.get('tables'):
        # AI Bank Parser result - same as UI
        tables_to_export = result_data.get('tables', [])
    elif 'forms_data' in result_data and result_data.get('forms_data'):
        # Form Data Parser result - convert forms to tables
        forms_data = result_data.get('forms_data', [])
        for page_data in forms_data:
            headers = page_data.get('headers', [])
            rows = page_data.get('data', [])
            if headers and rows:
                tables_to_export.append({
                    'table_id': len(tables_to_export) + 1,
                    'headers': headers,
                    'rows': rows,
                    'page_number': page_data.get('page_number', 0)
                })
    elif 'individual_results' in result_data:
        # Bulk job result - combine all tables from individual results
        for individual_result in result_data.get('individual_results', []):
            if 'tables' in individual_result:
                tables_to_export.extend(individual_result['tables'])
            elif 'forms_data' in individual_result:
                forms_data = individual_result['forms_data']
                for page_data in forms_data:
                    headers = page_data.get('headers', [])
                    rows = page_data.get('data', [])
                    if headers and rows:
                        tables_to_export.append({
                            'table_id': len(tables_to_export) + 1,
                            'headers': headers,
                            'rows': rows,
                            'page_number': page_data.get('page_number', 0)
                        })
    elif 'transactions' in result_data:
        # TEMPORARY: Handle old structure while debugging
        transactions = result_data.get('transactions', [])
        if transactions:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bank Statement"
            
            fieldnames = set()
            for txn in transactions:
                fieldnames.update(txn.keys())
            fieldnames = sorted(list(fieldnames))
            
            # Add title
            file_name = result_data.get('file_name', 'Bank Statement')
            extraction_method = result_data.get('extraction_method', 'API')
            ws.cell(row=1, column=1, value=f"{file_name} - {extraction_method.upper()}")
            ws.cell(row=2, column=1, value=f"Exported: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            
            # Add headers
            for col_idx, fieldname in enumerate(fieldnames, 1):
                ws.cell(row=4, column=col_idx, value=fieldname)
            
            # Add data
            for row_idx, txn in enumerate(transactions, 5):
                for col_idx, fieldname in enumerate(fieldnames, 1):
                    ws.cell(row=row_idx, column=col_idx, value=txn.get(fieldname, ''))
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
    
    if not tables_to_export:
        # Return empty Excel with message if no data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "No Data"
        ws.cell(row=1, column=1, value=f"No data found for export. Structure: {list(result_data.keys())}")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    # Create ZIP with multiple Excel files (same as UI)
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for table_idx, table in enumerate(tables_to_export):
            table_id = table.get('table_id', table_idx + 1)
            page_number = table.get('page_number', 0) + 1  # Convert to 1-based
            headers = table.get('headers', [])
            rows = table.get('rows', [])
            
            if not headers or not rows:
                continue
                
            filename = f"table_{table_id}_page_{page_number}.xlsx"
            content = generate_excel_content(headers, rows)
            zip_file.writestr(filename, content)
    
    zip_buffer.seek(0)
    
    from fastapi.responses import Response
    return Response(
        content=zip_buffer.getvalue(),
        media_type='application/zip',
        headers={'Content-Disposition': 'attachment; filename=bank-statement-tables.zip'}
    )


def export_to_txt(result_data: Dict[str, Any]):
    """Convert job result to TXT format"""
    from fastapi.responses import Response
    import io
    from datetime import datetime
    
    # Convert tables to transactions (flatten table rows into individual transaction dicts)
    transactions = []
    
    # Handle 'tables' structure first (single job)
    if 'tables' in result_data and result_data.get('tables'):
        for table in result_data['tables']:
            headers = table.get('headers', [])
            rows = table.get('rows', [])
            for row in rows:
                if len(headers) == len(row):
                    transactions.append(dict(zip(headers, row)))
    
    # Handle bulk job results - extract tables from individual_results
    elif 'individual_results' in result_data:
        for individual_result in result_data.get('individual_results', []):
            # Each individual result should have 'tables'
            if 'tables' in individual_result:
                for table in individual_result['tables']:
                    headers = table.get('headers', [])
                    rows = table.get('rows', [])
                    for row in rows:
                        if len(headers) == len(row):
                            transactions.append(dict(zip(headers, row)))
            # Also handle forms_data if present
            elif 'forms_data' in individual_result:
                for page_data in individual_result['forms_data']:
                    headers = page_data.get('headers', [])
                    rows = page_data.get('data', [])
                    for row in rows:
                        if len(headers) == len(row):
                            transactions.append(dict(zip(headers, row)))
    
    # Fallback to old logic if no tables found
    elif 'transactions' in result_data:
        # Single job result (old format)
        transactions = result_data.get('transactions', [])
    elif 'result' in result_data and isinstance(result_data['result'], dict):
        # Nested result structure
        transactions = result_data['result'].get('transactions', [])
    
    if not transactions:
        return Response(
            content=f"No tables or data found for export. Structure: {list(result_data.keys())}",
            media_type='text/plain'
        )
    
    output = io.StringIO()
    
    # Header
    if 'individual_results' in result_data:
        # Bulk job - use bulk info
        file_name = f"Bulk Export ({result_data.get('total_files', 0)} files)"
        extraction_method = "Bulk API"
    else:
        # Single job
        file_name = result_data.get('file_name', 'Bank Statement')
        extraction_method = result_data.get('extraction_method', 'API')
    
    output.write(f'Bank Statement Export - {extraction_method.upper()}\n')
    output.write('=' * 60 + '\n\n')
    output.write(f'File: {file_name}\n')
    output.write(f'Exported: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}\n')
    output.write(f'Total Transactions: {len(transactions)}\n\n')
    
    # Get all unique field names
    fieldnames = set()
    for txn in transactions:
        fieldnames.update(txn.keys())
    fieldnames = sorted(list(fieldnames))
    
    # Write header
    header_line = ' | '.join([field.ljust(15) for field in fieldnames])
    output.write(header_line + '\n')
    output.write('-' * len(header_line) + '\n')
    
    # Write transactions
    for txn in transactions:
        values = []
        for field in fieldnames:
            value = str(txn.get(field, '')).strip()
            values.append(value[:15].ljust(15))  # Truncate and pad
        
        output.write(' | '.join(values) + '\n')
    
    # Summary
    output.write('\n' + '=' * 60 + '\n')
    output.write(f'Summary: {len(transactions)} transactions exported\n')
    
    txt_content = output.getvalue()
    output.close()
    
    return Response(
        content=txt_content,
        media_type='text/plain',
        headers={'Content-Disposition': 'attachment; filename=bank_statement.txt'}
    )