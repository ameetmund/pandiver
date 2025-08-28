"""
Azure Document Intelligence Endpoints
Provides Intelligent Table Parser and Smart Key-Value Parser functionality
Replicates AWS Textract endpoints using Azure Document Intelligence Layout model
"""

from fastapi import APIRouter, File, UploadFile, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import tempfile
import os
import uuid
import pandas as pd
import io
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from datetime import datetime
import csv
import json
import re
from .azure_document_intelligence import (
    start_layout_analysis,
    get_analysis_status,
    extract_tables_from_azure_result,
    extract_key_value_pairs_from_azure_result,
    cleanup_operation
)

router = APIRouter()

# Response models (matching AWS Textract structure)
class AzureDIJobResponse(BaseModel):
    job_id: str
    status: str
    message: str

class AzureDIJobStatus(BaseModel):
    job_id: str
    status: str
    error: Optional[str] = None
    tables: Optional[List[Dict]] = None

class ExtractedTable(BaseModel):
    table_id: int
    headers: List[str]
    rows: List[List[str]]
    page_number: int

class MergeTablesRequest(BaseModel):
    tables: List[Dict[str, Any]]

class MergedDataResponse(BaseModel):
    headers: List[str]
    rows: List[List[str]]
    total_rows: int
    source_tables: int


# ============================
# INTELLIGENT TABLE PARSER ENDPOINTS
# ============================

@router.post("/intelligent-tables/start-analysis", response_model=AzureDIJobResponse)
async def start_intelligent_table_analysis(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...)
):
    """Start asynchronous Azure Document Intelligence analysis for intelligent table extraction"""
    
    if file.content_type != 'application/pdf':
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
    
    try:
        # Read file content
        file_content = await file.read()
        
        # Start Azure DI analysis
        operation_id = start_layout_analysis(file_content, file.filename)
        
        return AzureDIJobResponse(
            job_id=operation_id,
            status='IN_PROGRESS',
            message='Azure Document Intelligence table analysis started successfully'
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to start intelligent table analysis: {str(e)}")


@router.get("/intelligent-tables/job-status/{job_id}", response_model=AzureDIJobStatus)
async def get_intelligent_table_job_status(job_id: str):
    """Get the status of an Azure Document Intelligence table analysis job"""
    
    try:
        status_info = get_analysis_status(job_id)
        
        return AzureDIJobStatus(
            job_id=job_id,
            status=status_info['status'],
            error=status_info.get('error'),
            tables=None  # Tables will be available through process-results endpoint
        )
        
    except HTTPException:
        raise
    except Exception as e:
        return AzureDIJobStatus(
            job_id=job_id,
            status='FAILED',
            error=str(e)
        )


@router.get("/intelligent-tables/process-results/{job_id}")
async def process_intelligent_table_results(job_id: str):
    """Process and extract tables from completed Azure Document Intelligence analysis"""
    
    try:
        # Extract tables from Azure DI result
        tables = extract_tables_from_azure_result(job_id)
        
        return {
            'success': True,
            'tables': tables,
            'total_tables': len(tables)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process intelligent table results: {str(e)}")


@router.post("/intelligent-tables/merge-tables", response_model=MergedDataResponse)
async def merge_intelligent_tables(request: MergeTablesRequest):
    """Merge multiple extracted intelligent tables into a single dataset"""
    
    if not request.tables:
        raise HTTPException(status_code=400, detail="No tables provided for merging")
    
    try:
        # Simply concatenate all tables sequentially with spacing
        merged_rows = []
        merged_headers = []
        max_columns = 0
        
        # First pass: determine the maximum number of columns needed
        for table in request.tables:
            table_headers = table.get('headers', [])
            table_rows = table.get('rows', [])
            
            # Calculate max columns for this table (headers + data)
            current_max = len(table_headers)
            for row in table_rows:
                current_max = max(current_max, len(row))
            
            max_columns = max(max_columns, current_max)
        
        # Create generic headers for the maximum columns
        merged_headers = [f'Column {i+1}' for i in range(max_columns)]
        
        # Second pass: merge all tables sequentially
        for table_idx, table in enumerate(request.tables):
            table_headers = table.get('headers', [])
            table_rows = table.get('rows', [])
            
            # Add spacing between tables (except for the first one)
            if table_idx > 0:
                # Add 2 empty rows for spacing
                empty_row = [''] * max_columns
                merged_rows.append(empty_row)
                merged_rows.append(empty_row)
            
            # Add table identifier row
            table_identifier_row = [''] * max_columns
            table_identifier_row[0] = f'--- Table {table.get("table_id", table_idx + 1)} - Page {table.get("page_number", 0) + 1} ---'
            merged_rows.append(table_identifier_row)
            
            # Add one empty row
            empty_row = [''] * max_columns
            merged_rows.append(empty_row)
            
            # Add headers if they exist
            if table_headers:
                header_row = [''] * max_columns
                for i, header in enumerate(table_headers[:max_columns]):
                    header_row[i] = header or ''
                merged_rows.append(header_row)
            
            # Add all table rows as they are
            for row in table_rows:
                merged_row = [''] * max_columns
                for i, cell_value in enumerate(row[:max_columns]):
                    merged_row[i] = cell_value or ''
                merged_rows.append(merged_row)
        
        return MergedDataResponse(
            headers=merged_headers,
            rows=merged_rows,
            total_rows=len(merged_rows),
            source_tables=len(request.tables)
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to merge intelligent tables: {str(e)}")


@router.post("/intelligent-tables/export/{format}")
async def export_intelligent_table_data(format: str, data: dict):
    """Export intelligent table data in various formats"""
    
    supported_formats = ['csv', 'xlsx', 'json', 'xml', 'txt']
    if format not in supported_formats:
        raise HTTPException(status_code=400, detail=f"Unsupported format. Supported: {supported_formats}")
    
    # Check if this is multiple tables or merged data
    tables = data.get('tables', [])
    
    if tables:
        # Multiple tables - create separate files
        return export_multiple_intelligent_tables(format, tables)
    else:
        # Single merged data
        headers = data.get('headers', [])
        rows = data.get('rows', [])
        
        if not headers or not rows:
            raise HTTPException(status_code=400, detail="No data provided for export")
        
        try:
            if format == 'csv':
                return export_csv(headers, rows, "intelligent-tables")
            elif format == 'xlsx':
                return export_excel(headers, rows, "intelligent-tables")
            elif format == 'json':
                return export_json(headers, rows, data, "azure_intelligent_tables")
            elif format == 'xml':
                return export_xml(headers, rows, data, "azure_intelligent_tables")
            elif format == 'txt':
                return export_txt(headers, rows, data, "Azure Intelligent Tables")
                
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to export {format}: {str(e)}")


# ============================
# SMART KEY-VALUE PARSER ENDPOINTS
# ============================

@router.post("/smart-key-value/start-analysis", response_model=AzureDIJobResponse)
async def start_smart_key_value_analysis(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...)
):
    """Start asynchronous Azure Document Intelligence analysis for smart key-value extraction"""
    
    if file.content_type != 'application/pdf':
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
    
    try:
        # Read file content
        file_content = await file.read()
        
        # Start Azure DI analysis (same Layout model extracts both tables and key-value pairs)
        operation_id = start_layout_analysis(file_content, file.filename)
        
        return AzureDIJobResponse(
            job_id=operation_id,
            status='IN_PROGRESS',
            message='Azure Document Intelligence key-value analysis started successfully'
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to start smart key-value analysis: {str(e)}")


@router.get("/smart-key-value/job-status/{job_id}", response_model=AzureDIJobStatus)
async def get_smart_key_value_job_status(job_id: str):
    """Get the status of an Azure Document Intelligence key-value analysis job"""
    
    try:
        status_info = get_analysis_status(job_id)
        
        return AzureDIJobStatus(
            job_id=job_id,
            status=status_info['status'],
            error=status_info.get('error'),
            tables=None  # Key-value pairs will be available through process-results endpoint
        )
        
    except HTTPException:
        raise
    except Exception as e:
        return AzureDIJobStatus(
            job_id=job_id,
            status='FAILED',
            error=str(e)
        )


@router.get("/smart-key-value/process-results/{job_id}")
async def process_smart_key_value_results(job_id: str):
    """Process and extract key-value pairs from completed Azure Document Intelligence analysis"""
    
    try:
        # Extract key-value pairs from Azure DI result
        forms_data = extract_key_value_pairs_from_azure_result(job_id)
        
        return {
            'success': True,
            'forms_data': forms_data,
            'total_pages': len(forms_data)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process smart key-value results: {str(e)}")


@router.post("/smart-key-value/export/{format}")
async def export_smart_key_value_data(format: str, data: dict):
    """Export smart key-value data in various formats"""
    
    supported_formats = ['csv', 'xlsx', 'json', 'txt']
    if format not in supported_formats:
        raise HTTPException(status_code=400, detail=f"Unsupported format. Supported: {supported_formats}")
    
    forms_data = data.get('forms_data', [])
    
    if not forms_data:
        raise HTTPException(status_code=400, detail="No forms data provided for export")
    
    try:
        if len(forms_data) > 1:
            # Multiple pages - create separate files in ZIP
            return export_multiple_key_value_pages(format, forms_data)
        else:
            # Single page
            page_data = forms_data[0]
            headers = page_data.get('headers', [])
            rows = page_data.get('data', [])
            
            if format == 'csv':
                return export_csv(headers, rows, "smart-key-value")
            elif format == 'xlsx':
                return export_excel(headers, rows, "smart-key-value")
            elif format == 'json':
                return export_key_value_json(forms_data)
            elif format == 'txt':
                return export_key_value_txt(forms_data)
                
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export {format}: {str(e)}")


# ============================
# UTILITY FUNCTIONS
# ============================

def export_multiple_intelligent_tables(format: str, tables: List[Dict]):
    """Export multiple intelligent tables as a zip file containing separate files for each table"""
    import zipfile
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for table_idx, table in enumerate(tables):
            table_id = table.get('table_id', table_idx + 1)
            page_number = table.get('page_number', 0) + 1  # Convert to 1-based
            headers = table.get('headers', [])
            rows = table.get('rows', [])
            
            if not headers or not rows:
                continue
                
            filename = f"intelligent_table_{table_id}_page_{page_number}"
            
            if format == 'csv':
                content = generate_csv_content(headers, rows)
                zip_file.writestr(f"{filename}.csv", content)
            elif format == 'xlsx':
                content = generate_excel_content(headers, rows)
                zip_file.writestr(f"{filename}.xlsx", content)
            elif format == 'json':
                content = generate_json_content(headers, rows, table, "azure_intelligent_table")
                zip_file.writestr(f"{filename}.json", content)
            elif format == 'xml':
                content = generate_xml_content(headers, rows, table)
                zip_file.writestr(f"{filename}.xml", content)
            elif format == 'txt':
                content = generate_txt_content(headers, rows, table)
                zip_file.writestr(f"{filename}.txt", content)
    
    zip_buffer.seek(0)
    
    return StreamingResponse(
        io.BytesIO(zip_buffer.getvalue()),
        media_type='application/zip',
        headers={'Content-Disposition': f'attachment; filename=intelligent-tables.zip'}
    )


def export_multiple_key_value_pages(format: str, forms_data: List[Dict]):
    """Export multiple key-value pages as a zip file containing separate files for each page"""
    import zipfile
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for page_data in forms_data:
            page_number = page_data.get('page_number', 0) + 1  # Convert to 1-based
            headers = page_data.get('headers', [])
            rows = page_data.get('data', [])
            
            if not headers:
                continue
                
            filename = f"smart_key_value_page_{page_number}"
            
            if format == 'csv':
                content = generate_csv_content(headers, rows)
                zip_file.writestr(f"{filename}.csv", content)
            elif format == 'xlsx':
                content = generate_excel_content(headers, rows)
                zip_file.writestr(f"{filename}.xlsx", content)
            elif format == 'json':
                content = generate_key_value_json_content(page_data)
                zip_file.writestr(f"{filename}.json", content)
            elif format == 'txt':
                content = generate_key_value_txt_content(page_data)
                zip_file.writestr(f"{filename}.txt", content)
    
    zip_buffer.seek(0)
    
    return StreamingResponse(
        io.BytesIO(zip_buffer.getvalue()),
        media_type='application/zip',
        headers={'Content-Disposition': f'attachment; filename=smart-key-value-pages.zip'}
    )


def generate_csv_content(headers, rows):
    """Generate CSV content as string"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


def generate_excel_content(headers, rows):
    """Generate Excel content as bytes"""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Table Data"
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, cell_value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    wb.save(output)
    return output.getvalue()


def generate_json_content(headers, rows, table_info, export_type):
    """Generate JSON content as string"""
    data = {
        'table_id': table_info.get('table_id', 1),
        'page_number': table_info.get('page_number', 0) + 1,
        'headers': headers,
        'data': [
            {headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))}
            for row in rows
        ],
        'metadata': {
            'total_rows': len(rows),
            'export_date': datetime.now().isoformat(),
            'export_type': export_type
        }
    }
    return json.dumps(data, indent=2, ensure_ascii=False)


def generate_xml_content(headers, rows, table_info):
    """Generate XML content as string"""
    root = ET.Element('table')
    
    # Metadata
    metadata = ET.SubElement(root, 'metadata')
    ET.SubElement(metadata, 'table_id').text = str(table_info.get('table_id', 1))
    ET.SubElement(metadata, 'page_number').text = str(table_info.get('page_number', 0) + 1)
    ET.SubElement(metadata, 'total_rows').text = str(len(rows))
    ET.SubElement(metadata, 'export_date').text = datetime.now().isoformat()
    
    # Data
    data_element = ET.SubElement(root, 'data')
    
    for idx, row in enumerate(rows):
        row_element = ET.SubElement(data_element, 'row', id=str(idx + 1))
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else ''
            clean_header = re.sub(r'[^a-zA-Z0-9]', '_', header.lower())
            ET.SubElement(row_element, clean_header).text = str(value)
    
    return ET.tostring(root, encoding='unicode')


def generate_txt_content(headers, rows, table_info):
    """Generate TXT content as string"""
    output = io.StringIO()
    
    output.write(f'Table {table_info.get("table_id", 1)} - Page {table_info.get("page_number", 0) + 1}\n')
    output.write('=' * 50 + '\n\n')
    
    # Headers
    output.write(' | '.join(headers) + '\n')
    output.write('-' * (len(' | '.join(headers))) + '\n')
    
    # Rows
    for row in rows:
        output.write(' | '.join(str(cell) for cell in row) + '\n')
    
    output.write(f'\nTotal Rows: {len(rows)}\n')
    
    return output.getvalue()


def generate_key_value_json_content(page_data):
    """Generate JSON content for key-value data"""
    data = {
        'page_number': page_data.get('page_number', 0) + 1,
        'headers': page_data.get('headers', []),
        'data': page_data.get('data', []),
        'key_value_pairs': page_data.get('key_value_pairs', []),
        'metadata': {
            'export_date': datetime.now().isoformat(),
            'export_type': 'azure_smart_key_value'
        }
    }
    return json.dumps(data, indent=2, ensure_ascii=False)


def generate_key_value_txt_content(page_data):
    """Generate TXT content for key-value data"""
    output = io.StringIO()
    
    page_number = page_data.get('page_number', 0) + 1
    headers = page_data.get('headers', [])
    rows = page_data.get('data', [])
    
    output.write(f'Smart Key-Value Data - Page {page_number}\n')
    output.write('=' * 50 + '\n\n')
    
    if headers and rows:
        # Headers
        output.write(' | '.join(headers) + '\n')
        output.write('-' * (len(' | '.join(headers))) + '\n')
        
        # Rows
        for row in rows:
            output.write(' | '.join(str(cell) for cell in row) + '\n')
    
    # Also include key-value pairs
    key_value_pairs = page_data.get('key_value_pairs', [])
    if key_value_pairs:
        output.write('\n\nKey-Value Pairs:\n')
        output.write('-' * 20 + '\n')
        for pair in key_value_pairs:
            output.write(f"{pair.get('key', '')}: {pair.get('value', '')}\n")
    
    return output.getvalue()


def export_csv(headers, rows, feature_name):
    """Export data as CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write headers
    writer.writerow(headers)
    
    # Write rows
    for row in rows:
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename={feature_name}-azure.csv'}
    )


def export_excel(headers, rows, feature_name):
    """Export data as Excel file"""
    output = io.BytesIO()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = feature_name.replace('-', ' ').title()
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, cell_value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={feature_name}-azure.xlsx'}
    )


def export_json(headers, rows, data, export_type):
    """Export data as JSON"""
    export_data = {
        'headers': headers,
        'transactions': [
            {headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))}
            for row in rows
        ],
        'metadata': {
            'total_transactions': len(rows),
            'source_tables': data.get('source_tables', 0),
            'export_date': datetime.now().isoformat(),
            'export_type': export_type
        }
    }
    
    json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
    
    return StreamingResponse(
        io.BytesIO(json_str.encode('utf-8')),
        media_type='application/json',
        headers={'Content-Disposition': f'attachment; filename={export_type}-azure.json'}
    )


def export_xml(headers, rows, data, export_type):
    """Export data as XML"""
    root = ET.Element('document')
    
    # Metadata
    metadata = ET.SubElement(root, 'metadata')
    ET.SubElement(metadata, 'total_transactions').text = str(len(rows))
    ET.SubElement(metadata, 'source_tables').text = str(data.get('source_tables', 0))
    ET.SubElement(metadata, 'export_date').text = datetime.now().isoformat()
    ET.SubElement(metadata, 'export_type').text = export_type
    
    # Transactions
    transactions = ET.SubElement(root, 'transactions')
    
    for idx, row in enumerate(rows):
        transaction = ET.SubElement(transactions, 'transaction', id=str(idx + 1))
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else ''
            clean_header = re.sub(r'[^a-zA-Z0-9]', '_', header.lower())
            ET.SubElement(transaction, clean_header).text = str(value)
    
    xml_str = ET.tostring(root, encoding='unicode')
    
    return StreamingResponse(
        io.BytesIO(xml_str.encode('utf-8')),
        media_type='application/xml',
        headers={'Content-Disposition': f'attachment; filename={export_type}-azure.xml'}
    )


def export_txt(headers, rows, data, title):
    """Export data as text file"""
    output = io.StringIO()
    
    # Header
    output.write(f'{title} - Azure Document Intelligence\n')
    output.write('=' * 50 + '\n\n')
    
    # Metadata
    output.write(f'Total Transactions: {len(rows)}\n')
    output.write(f'Source Tables: {data.get("source_tables", 0)}\n')
    output.write(f'Export Date: {datetime.now().isoformat()}\n\n')
    
    # Headers
    output.write(' | '.join(headers) + '\n')
    output.write('-' * (len(' | '.join(headers))) + '\n')
    
    # Rows
    for row in rows:
        output.write(' | '.join(str(cell) for cell in row) + '\n')
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type='text/plain',
        headers={'Content-Disposition': f'attachment; filename={title.lower().replace(" ", "-")}-azure.txt'}
    )


def export_key_value_json(forms_data):
    """Export key-value data as JSON"""
    export_data = {
        'forms_data': forms_data,
        'metadata': {
            'total_pages': len(forms_data),
            'export_date': datetime.now().isoformat(),
            'export_type': 'azure_smart_key_value'
        }
    }
    
    json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
    
    return StreamingResponse(
        io.BytesIO(json_str.encode('utf-8')),
        media_type='application/json',
        headers={'Content-Disposition': 'attachment; filename=smart-key-value-azure.json'}
    )


def export_key_value_txt(forms_data):
    """Export key-value data as text file"""
    output = io.StringIO()
    
    # Header
    output.write('Smart Key-Value Extraction - Azure Document Intelligence\n')
    output.write('=' * 50 + '\n\n')
    
    # Process each page
    for page_data in forms_data:
        page_number = page_data.get('page_number', 0) + 1
        headers = page_data.get('headers', [])
        rows = page_data.get('data', [])
        key_value_pairs = page_data.get('key_value_pairs', [])
        
        output.write(f'Page {page_number}\n')
        output.write('-' * 20 + '\n')
        
        if headers and rows:
            # Table format
            output.write(' | '.join(headers) + '\n')
            output.write('-' * (len(' | '.join(headers))) + '\n')
            for row in rows:
                output.write(' | '.join(str(cell) for cell in row) + '\n')
        
        # Key-value pairs
        if key_value_pairs:
            output.write('\nKey-Value Pairs:\n')
            for pair in key_value_pairs:
                output.write(f"{pair.get('key', '')}: {pair.get('value', '')}\n")
        
        output.write('\n')
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type='text/plain',
        headers={'Content-Disposition': 'attachment; filename=smart-key-value-azure.txt'}
    )



# ============================
# INTELLIGENT DATA PARSER ENDPOINTS (COMBINED)
# ============================

@router.post("/intelligent-data/start-analysis", response_model=AzureDIJobResponse)
async def start_intelligent_data_analysis(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...)
):
    """Start asynchronous Azure Document Intelligence analysis for combined table and key-value extraction"""
    
    if file.content_type != 'application/pdf':
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
    
    try:
        # Read file content
        file_content = await file.read()
        
        # Start Azure DI analysis (single call for both tables and key-value pairs)
        operation_id = start_layout_analysis(file_content, file.filename)
        
        return AzureDIJobResponse(
            job_id=operation_id,
            status='IN_PROGRESS',
            message='Azure Document Intelligence combined analysis started successfully'
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to start intelligent data analysis: {str(e)}")


@router.get("/intelligent-data/job-status/{job_id}", response_model=AzureDIJobStatus)
async def get_intelligent_data_job_status(job_id: str):
    """Get the status of an Azure Document Intelligence combined analysis job"""
    
    try:
        status_info = get_analysis_status(job_id)
        
        return AzureDIJobStatus(
            job_id=job_id,
            status=status_info['status'],
            error=status_info.get('error'),
            tables=None  # Combined data will be available through process-results endpoint
        )
        
    except HTTPException:
        raise
    except Exception as e:
        return AzureDIJobStatus(
            job_id=job_id,
            status='FAILED',
            error=str(e)
        )


@router.get("/intelligent-data/process-results/{job_id}")
async def process_intelligent_data_results(job_id: str):
    """Process and extract both tables and key-value pairs from completed Azure Document Intelligence analysis"""
    
    try:
        # Extract both tables and key-value pairs from the same Azure DI result
        tables = extract_tables_from_azure_result(job_id)
        forms_data = extract_key_value_pairs_from_azure_result(job_id)
        
        return {
            'success': True,
            'tables': tables,
            'total_tables': len(tables),
            'forms_data': forms_data,
            'total_pages': len(forms_data)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process intelligent data results: {str(e)}")


@router.post("/intelligent-data/export-tables/{format}")
async def export_intelligent_data_tables(format: str, data: dict):
    """Export table data from intelligent data parser"""
    
    supported_formats = ['csv', 'xlsx', 'json', 'xml', 'txt']
    if format not in supported_formats:
        raise HTTPException(status_code=400, detail=f"Unsupported format. Supported: {supported_formats}")
    
    # Check if this is multiple tables or merged data
    tables = data.get('tables', [])
    
    if tables:
        # Multiple tables - create separate files
        return export_multiple_intelligent_tables(format, tables)
    else:
        # Single merged data
        headers = data.get('headers', [])
        rows = data.get('rows', [])
        
        if not headers or not rows:
            raise HTTPException(status_code=400, detail="No data provided for export")
        
        try:
            if format == 'csv':
                return export_csv(headers, rows, "intelligent-data-tables")
            elif format == 'xlsx':
                return export_excel(headers, rows, "intelligent-data-tables")
            elif format == 'json':
                return export_json(headers, rows, data, "azure_intelligent_data_tables")
            elif format == 'xml':
                return export_xml(headers, rows, data, "azure_intelligent_data_tables")
            elif format == 'txt':
                return export_txt(headers, rows, data, "Azure Intelligent Data Tables")
                
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to export {format}: {str(e)}")


@router.post("/intelligent-data/export-key-values/{format}")
async def export_intelligent_data_key_values(format: str, data: dict):
    """Export key-value data from intelligent data parser"""
    
    supported_formats = ['csv', 'xlsx', 'json', 'txt']
    if format not in supported_formats:
        raise HTTPException(status_code=400, detail=f"Unsupported format. Supported: {supported_formats}")
    
    forms_data = data.get('forms_data', [])
    
    if not forms_data:
        raise HTTPException(status_code=400, detail="No forms data provided for export")
    
    try:
        if len(forms_data) > 1:
            # Multiple pages - create separate files in ZIP
            return export_multiple_key_value_pages(format, forms_data)
        else:
            # Single page
            page_data = forms_data[0]
            headers = page_data.get('headers', [])
            rows = page_data.get('data', [])
            
            if format == 'csv':
                return export_csv(headers, rows, "intelligent-data-key-values")
            elif format == 'xlsx':
                return export_excel(headers, rows, "intelligent-data-key-values")
            elif format == 'json':
                return export_key_value_json(forms_data)
            elif format == 'txt':
                return export_key_value_txt(forms_data)
                
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export {format}: {str(e)}")


# Cleanup endpoint to remove old operations
@router.delete("/cleanup-operations")
async def cleanup_old_operations():
    """Remove operations older than 24 hours"""
    
    from .azure_document_intelligence import azure_di_jobs
    
    current_time = datetime.now()
    operations_to_remove = []
    
    for operation_id, job_info in azure_di_jobs.items():
        started_at = datetime.fromisoformat(job_info['started_at'])
        if (current_time - started_at).total_seconds() > 86400:  # 24 hours
            operations_to_remove.append(operation_id)
    
    for operation_id in operations_to_remove:
        cleanup_operation(operation_id)
    
    return {
        'message': f'Cleaned up {len(operations_to_remove)} old operations',
        'removed_operations': len(operations_to_remove)
    }